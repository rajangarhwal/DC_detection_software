import Tkinter, Tkconstants, tkFileDialog
from Tkinter import *
from ranking import rank, find_good_metaphases
import os
import matplotlib.pyplot as plt
from PIL import Image, ImageTk, ImageGrab
from os.path import basename
import tkMessageBox
import shutil
import ttk
import threading
import numpy as np
import openpyxl
import time
import cv2

segment_path="C:/Users/Acer/Desktop/1images/segments/scoredData.xlsx"
wb = openpyxl.load_workbook(segment_path)
ws = wb.active
col=ws.max_column
print(col)    
c1=ws.cell(1, col)
c1.value="Dicentric"
c1=ws.cell(1, col+1)
c1.value="Monocentric"
c1=ws.cell(1, col+2)
c1.value="Total Chromsome"
c1=ws.cell(2, col)
if c1.value is None:
    c1.value="Tu cheutia hai"
    print("tu chutia hai")
else:
    print("well done")
wb.save(segment_path)