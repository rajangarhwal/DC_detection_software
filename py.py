import numpy as np

f=open("file.txt",'w')
f.write("sd fdfdhsfjds\n")
i=4
f.write('{}  {}\n'.format(i+1,i+2))
f.write('{}  {}\n'.format(i+1,i+2))
f.write('{}  {}\n'.format(i+1,i+2))
f.close()
data = np.genfromtxt('file.txt')
print(data)
row,col=np.shape(data)
print(row,col,row+col+1)    




'''
def retrieve():
        print("sd;fkjdb shjfbdjsh fbjhdsbfhjdsbfhjdbshjf bdshj fdshjvfhjdsvhfvdshf sdhjfjha")
        f =open(undo_path,'r')
        f.seek(0, 0)
        data = np.genfromtxt(undo_path)
        f.close()
        data=np.delete(data,(0),axis=0)
        f.write("x1  y1  x2  y2  colour  id\n")
        f =open(undo_path,'w')
        for z in data:
            x1,y1,x2,y2,colour,id=z
            print(x1,y1,x2,y2,colour,id)
            if(colour==1):
                id=canvas.create_oval(x1,y1,x2,y2,fill='',width=2,outline='green')
            else:
                id=canvas.create_oval(x1,y1,x2,y2,fill='',width=2,outline='red')
            f.write('{}  {}  {}  {}  {} {}\n'.format(x1,y1,x2,y2,colour,id))
            f.close()
            f =open(redo_path,'w')
            f.close()







            
import os
import pandas as pd
#C:\Users\Acer\Desktop\p images\results_NB7_A.39
path='C:\Users\Acer\Desktop\p images\segments\scoredData_1565692952.2.xlsx'
print(path)
import openpyxl

wb = openpyxl.load_workbook(path) 
sheet = wb.active 
  
max_col = sheet.max_column 
print(sheet.max_row)
print(sheet.max_column)
c1=sheet.cell(row=4,column=4)
c1.value="you have done it"  
# Will print a particular row value 
for i in range(1, max_col + 1): 
    cell_obj = sheet.cell(row = 2, column = i) 
    print(cell_obj.value,"  ")
wb.save(path) '''

'''
import Tkinter, Tkconstants, tkFileDialog
from Tkinter import *
from ranking import rank, find_good_metaphases
import os
import matplotlib.pyplot as plt
from PIL import Image, ImageTk, ImageGrab
from os.path import basename
import tkMessageBox
import shutil
import ttk
import threading
import numpy as np
import openpyxl
import time
import cv2
#import io
selected_file = ""
selected_folder = ""
pathss=""

dirtext='Select a Directory'
filetext= 'Select a file'

initial_directory = ""

class Metaphase_Analysis(Tkinter.Frame):
    def __init__(self, root):
        Tkinter.Frame.__init__(self, root)

        button_opt = {'fill': Tkconstants.BOTH, 'padx': 5, 'pady': 5}
        # self.file_button = Tkinter.Button(self, text = filetext, command = self.openfile, height = 6, width = 20)
        # self.file_button.pack(side=LEFT, **button_opt)
        self.dir_button = Tkinter.Button(self, text = dirtext,font=(None, 12), command = self.opendirectory, height = 5, width = 50)
        self.dir_button.pack(side=LEFT, **button_opt)
        # self.refine_button = Tkinter.Button(self, text = "Refine", command = self.refine)
        # self.refine_button.pack(side=LEFT, **button_opt)
        self.exit_button = Tkinter.Button(self, text = "Exit",font=(None, 12), command = root.destroy, height = 5, width = 20)
        self.exit_button.pack(side=LEFT, **button_opt)
        self.help_button = Tkinter.Button(self, text = "Help",font=(None, 12), command = self.helptext, height = 5, width = 20)
        self.help_button.pack(side=LEFT, **button_opt)
        # self.chr_label = Tkinter.Label(self,text="Chromosome segments", height = 6, width=25)
        # self.chr_label.pack(side = RIGHT, **button_opt)
        self.metaphase_label = Tkinter.Label(self,text="Metaphases", font=(None, 12), height = 5, width=25)
        self.metaphase_label.pack(side = LEFT, padx = 40)
        #self.pack(fill=BOTH, expand=NO)
        self.switch_button = Tkinter.Button(self, text = "Switch",font=(None, 12), command = self.helptext, height = 5, width = 25)
        self.switch_button.pack(side=LEFT, **button_opt)

        # define options for opening or saving a file
        self.file_opt = options = {}
        global extension
        extension = '.JPG'
        options['defaultextension'] = extension
        options['filetypes'] = [('image files', extension)]
        options['initialdir'] = initial_directory
        options['initialfile'] = ''
        options['parent'] = root
        options['title'] = 'Metaphase Analysis'

        # defining options for opening a directory
        self.dir_opt = options = {}
        options['initialdir'] = initial_directory
        options['mustexist'] = False
        options['parent'] = root
        options['title'] = 'MetaPhase Analysis'

    def helptext(self):
        #textmsg = '1. text to display\n2. text to display\n3. text to display\n4. text to display\n5. text to display'
        f1 = open("Help.txt", "r")
        textmsg = f1.read()
        #print textmsg
        showdialog = tkMessageBox.showinfo('Help', textmsg)
        f1.close()

    def openfile(self):
        filename = tkFileDialog.askopenfilename(**self.file_opt)
        self.file_button["text"] = str(filename) if filename else filetext
        if filename:
            global selected_file
            selected_file = str(filename)
        if self.file_button["text"] == str(filename):
            self.quit()

    def opendirectory(self):
        dirname = tkFileDialog.askdirectory(**self.dir_opt)
        self.dir_button["text"] = str(dirname) if dirname else dirtext
        if dirname:
            global selected_folder
            selected_folder = str(dirname)
        if self.dir_button["text"] == str(dirname):
            self.quit()


# Adding scrollbar to the canvas
class VerticalScrolledFrame(Tkinter.Frame):
    def __init__(self, parent, *args, **kw):
        Tkinter.Frame.__init__(self, parent, *args, **kw)            

        vscrollbar = Tkinter.Scrollbar(self, orient=Tkinter.VERTICAL)
        vscrollbar.pack(fill=Tkinter.Y, side=Tkinter.LEFT, expand=Tkinter.FALSE)
        canvas = Tkinter.Canvas(self, bd=0, highlightthickness=0, yscrollcommand=vscrollbar.set)
        canvas.pack(side=Tkinter.LEFT, fill=Tkinter.BOTH, expand=Tkinter.TRUE)
        vscrollbar.config(command=canvas.yview)

        # reset the view
        canvas.xview_moveto(0)
        canvas.yview_moveto(0)
        canvas.delete("All")
        # create a frame inside the canvas which will be scrolled with it
        self.interior = interior = Tkinter.Frame(canvas)
        interior_id = canvas.create_window(0, 0, window=interior, anchor=Tkinter.NW)
        self.pack(fill=BOTH, expand=YES)


        def myfunction(event):
            canvas.configure(scrollregion=canvas.bbox("all"),width=1240,height=400 )

        interior.bind('<Configure>', myfunction)

        def _configure_canvas(event):
            if interior.winfo_reqwidth() != canvas.winfo_width():
                # update the inner frame's width to fill the canvas
                canvas.itemconfigure(interior_id, width=canvas.winfo_width())
        canvas.bind('<Configure>', _configure_canvas)


def make_gui_before():
    root = Tkinter.Tk()
    global downloaded
    downloaded = IntVar()
    root.attributes('-fullscreen', True)

    root.title("Detection of dicentric chromosomes")
    w, h = root.winfo_screenwidth(), root.winfo_screenheight()
    posx  = 0
    posy  = 0
    root.wm_geometry("%dx%d+%d+%d" % (w, h, posx, posy))

    statusbar = Tkinter.Label(root, borderwidth=1, relief="sunken")
    statusbar.pack(side="bottom", fill="x")
    
    Metaphase_Analysis(root).pack()
    f2 = Frame(root, bg = "black", height=h/1.5, width = w/2)
    f2.pack(side=LEFT, fill = Y)
    f3 = Frame(root, bg = "black", height=h/1.5, width = w/9)
    f3.pack(side=LEFT, fill = Y)
    f4 = Frame(root, bg = "black", height=0, width = 0)
    f4.pack(side=RIGHT, fill = Y)


    frm1 = Frame(root,bg = "black", height = h/1.3,width = w/50)
    frm1.pack(side = RIGHT, fill=Y)
    frm2 = Frame(root,bg = "black", height = h/1.3,width = w/50)
    frm2.pack(side = RIGHT, fill=Y)

    frm3 = Frame(root,bg = "black", height = h/1.3,width = w/20)
    frm3.pack(side = RIGHT, fill=Y)
    frm4 = Frame(root,bg = "black", height = h/1.3,width = w/50)
    frm4.pack(side = RIGHT, fill=Y)
    frm5 = Frame(root,bg = "black", height = h/1.3,width = w/50)
    frm5.pack(side = RIGHT, fill=Y)
    frm6 = Frame(root,bg = "black", height = h/1.3,width = w/50)
    frm6.pack(side = RIGHT, fill=Y)

    scframe = VerticalScrolledFrame(root)
    scframe.pack()

    canvas = Canvas(f2, width = w/2, height = h/2)
    canvas.pack(fill=X, expand=2)
    canvas2 = Canvas(f4, width = 0, height = 0)
    canvas2.pack(fill=X, expand=2)

    return root, f2, f3, scframe, canvas ,f4,  canvas2, frm1,frm2,frm3,frm4,frm5,frm6

class Threader(threading.Thread):

    def __init__(self, *args, **kwargs):

        threading.Thread.__init__(self, *args, **kwargs)
        self.daemon = True
        self.start()

    def run(self):
        root.quit()
        global originalPath,scoredImages,scoredImagesList, pres_time, rankPlot
        global mc_and_dc_list, number_of_segments, good_metaphases_list, good_metaphases_list_with_coordinates, file_with_good_metaphases, actual_contours_path, dict_of_coordinates
        mc_and_dc_list, number_of_segments,good_metaphases_list, good_metaphases_list_with_coordinates,segpath, dict_of_coordinates = find_good_metaphases(downloaded, progress,file_name_list,f2,root,25)
        originalPath,scoredImages, pres_time, scoredImagesList, rankPlot = rank(segpath)
        time.sleep(0.3)
        for i in range(0, len(originalPath)):
            #originalPath[i][0]
            rnk = originalPath[i][1][0]
            print "**********",originalPath[i][0], rnk, str(rnk)
            q = originalPath[i][0].split('/')
            imgname = q[len(q) - 1]
            segmented_img = ''
            for j in range(0, len(q)-1):
                segmented_img = segmented_img + '/' + str(q[j])
            segmented_img = segmented_img + '/segments/' + imgname
            img1 = cv2.imread(segmented_img)
            print segmented_img
            #print filename[-2]
            path_to_store = selected_folder + '/segments/score' + str(rnk)+ '/' + str(imgname)
            print path_to_store
            print "^^^", time.time()
            status = cv2.imwrite(path_to_store, img1)


class disp_segment:

    def __init__(self,root,image,i):
        if 0<=i<15:
            self.c = Canvas(frm1, height=60, width = 60)
            self.c.pack(fill=Y,expand=0)
            self.c.pack(fill=Y,expand=0)
            self.c.create_image(0,0,anchor = NW, image = image)
            self.c.image = image
        if 15<=i<30:
            self.c = Canvas(frm2, height=60, width = 60)
            self.c.pack(fill=Y,expand=0)
            self.c.pack(fill=Y,expand=0)
            self.c.create_image(0,0,anchor = NW, image = image)
            self.c.image = image
        if 30<=i<45:
            self.c = Canvas(frm3, height=60, width = 60)
            self.c.pack(fill=Y,expand=0)
            self.c.create_image(0,0,anchor = NW, image = image)
            self.c.image = image
        if 45<=i<60:
            self.c = Canvas(frm4, height=60, width = 60)
            self.c.pack(fill=Y,expand=0)
            self.c.pack(fill=Y,expand=0)
            self.c.create_image(0,0,anchor = NW, image = image)
            self.c.image = image
        if 60<=i<75:
            self.c = Canvas(frm5, height=60, width = 60)
            self.c.pack(fill=Y,expand=0)
            self.c.pack(fill=Y,expand=0)
            self.c.create_image(0,0,anchor = NW, image = image)
            self.c.image = image
        if 75<=i<90:
            self.c = Canvas(frm6, height=60, width = 60)
            self.c.pack(fill=Y,expand=0)
            self.c.pack(fill=Y,expand=0)
            self.c.create_image(0,0,anchor = NW, image = image)
            self.c.image = image
def show_segments(histPath):
    photos = []
    destroy(root)
    i=0
    print histPath
    for file in os.listdir(histPath):
        if file.endswith('.jpg'):
            img = ImageTk.PhotoImage(Image.open(histPath+file))  # PIL solution

            i+=1
def destroy(root):
    for widget in frm1.winfo_children():
        widget.destroy()
    for widget in frm2.winfo_children():
        widget.destroy()
    for widget in frm3.winfo_children():
        widget.destroy()
    for widget in frm4.winfo_children():
        widget.destroy()
    for widget in frm5.winfo_children():
        widget.destroy()
    for widget in frm6.winfo_children():
        widget.destroy()
def delete_dialogue_box(text):
    msg = "Are you sure to delete " + os.path.basename(text) + " ?"
    result = tkMessageBox.askquestion("Delete", msg, icon='warning')
    return result
def detect_dc(event,segments,text,extra=None):
    global single
    if extra == 1:
        single = True

        click('show dc', event, None,segments, text)
    return
def removekey(d, key):
    r = dict(d)
    if key in r :
        del r[key]
    return r

filename_ext = 'a'
prev=0
def click(msg, event, _btn, segments, text):

    global path, segment_path
    if msg == 'single click':
        dir = os.path.dirname(text)
        global dc_path 
        segment_path = dir+'/segments/'
        dc_path = segment_path +os.path.basename(text)
        global file_name1
        file_name1=os.path.basename(text)
        print file_name1,msg
        global histPath, filename_ext
        filename_ext = os.path.splitext(basename(text))[0]
        histPath = dir + "/results_" + filename_ext + '/actual/'
        show_segments(histPath)
        canvas.delete("all")
        path = text
        canvas.pack(fill=Tkinter.BOTH, expand=2) # expand
        image = Image.open(dir+"/"+os.path.basename(text))       
        width, height = image.size
        topleft = (600,500)
        bottomright = (2000,1600)
        cropped = image.crop((topleft[0], topleft[1], bottomright[0], bottomright[1]))
        width = bottomright[0] - topleft[0]
        height = bottomright[1] - topleft[1]
        image = cropped.resize((int(width/1.8), int(height/1.8)), Image.ANTIALIAS)
        #global photo1
        photo = ImageTk.PhotoImage(image)
        canvas.create_image(0, 0, image = photo, anchor = NW)
        canvas.photo = photo

        image_name_entry.delete(0, 'end')
        image_name_entry.insert(0, filename_ext)
        total_segments.delete(0, 'end')
        for i in range(0, len(scoredImages)):
            if scoredImages[i][0] == filename_ext:
                temp_i = i
                break
        total_segments.insert(0, str(scoredImages[temp_i][1][1]))
        metaphase_rank.delete(0, 'end')
        #print scoredImages[filename_ext][0]
        metaphase_rank.insert(0,str(scoredImages[temp_i][1][0]))
        
    elif msg == 'double click':
        images=[]
        #buttons=[]
        result = delete_dialogue_box(text)
        if result == 'yes':
            global prev
            prev = 2
            image = Image.open(text)
            image.paste(red_cross, (100, 100), red_cross)
            image = image.resize((100,100), Image.ANTIALIAS)
            photo = ImageTk.PhotoImage(image)
            event.widget.config(image=photo,width="100",height="100")
            images.append(photo)
            print "DELETING " + os.path.basename(text)
            global good_metaphases_list, good_metaphases_list_with_coordinates
            good_metaphases_list = removekey(good_metaphases_list, text)
            good_metaphases_list_with_coordinates = removekey(good_metaphases_list_with_coordinates, text) 
            dir = os.path.dirname(text)
            #filename_ext = os.path.splitext(basename(text))[0]
            #print "filename_ext", filename_ext
            newpath = dir + "/results_" + filename_ext
            shutil.rmtree(newpath)

    elif msg == 'show dc':
        dir = os.path.dirname(text)
        path = segment_path+"/"+os.path.basename(text)
        print dc_path, msg
        canvas.delete("all")

        canvas.pack(fill=Tkinter.BOTH, expand=2) # expand
        image = Image.open(dc_path)
        print text
        path = text
        #filename_ext = os.path.splitext(basename(text))[0]
        histPath = dir + "/results_" + filename_ext + '/actual/'
        width, height = image.size
        topleft = (600,500)
        bottomright = (2000,1600)
        cropped = image.crop((topleft[0], topleft[1], bottomright[0], bottomright[1]))
        width = bottomright[0] - topleft[0]
        height = bottomright[1] - topleft[1]
        image = cropped.resize((int(width/1.8), int(height/1.8)), Image.ANTIALIAS)
        global photo1
        photo1 = ImageTk.PhotoImage(image)
        canvas.create_image(0, 0, image = photo1, anchor = NW)
        canvas.photo = photo1

    try:
        return histPath
    except Exception as NameError:
        print "nothing to return"


def test(event,sorted_pathlist,buttons, segments, text, btnclicked, extra=None):
    global single
    if extra == 1 and btnclicked == 0:
        single = True
        root.after(500, single_click, event, segments, text)
        highlightimg(event, sorted_pathlist,buttons, segments, text, btnclicked)
    if extra == 1 and (btnclicked == 1 or btnclicked == 2):
        single = True
        #root.after(500, single_click, event, segments, textpre)
        highlightimg(event, sorted_pathlist,buttons, segments, text, btnclicked)
    elif extra == 101:
        single = False
        click('double click', event,sorted_pathlist, segments, text)


def single_click(event, segments, text):
    global single
    if single:
        single = False

        click('single click', event, None,segments, text)


#i_val=-1
def highlightimg(event, sorted_pathlist, btns, segments, text, btnclicked):
    global prev, i_val,pathss
    if btnclicked == 0:
        for i in range(0, len(sorted_pathlist)):
            if sorted_pathlist[i][0] == text:
                print("sdkfkjsdhfjds fdjksfjkdfkjds jfbdsjdfbsjdfb jds bfjsdjfds fdsf dsjf ds fhjds fjd")
                pathss=text[:-5]
                head,tail=os.path.split(pathss)
                pathss=head +'/results_'+tail+'/file.txt'
                f =open(pathss,'w')
                f.close
                print(pathss)
                if event and prev != 2:
                    btns[i].config(bg="red2")
                    if prev == 1:
                        btns[i_val].config(bg="gray99", fg="purple3", font="Dosis")
                    i_val = i
                if event and prev == 2:
                    btns[i].config(bg="red2")
                    i_val = i
                if i_val == i:
                    btns[i].config(bg="red2")
                    i_val = i
                break
    elif btnclicked == 1:
        if i_val <= 0:
            i_val = 0
            return
        if event:
            i_val -= 1
            n1 = 0
            if sorted_pathlist[i_val] == 'deleted':
                for h in range(1, i_val+1):
                    if sorted_pathlist[i_val-h] != 'deleted':
                        n1 = h
                        break
                i_val -= n1
                btns[i_val].config(bg="red2")
                btns[i_val+n1+1].config(bg="gray99", fg="purple3", font="Dosis")
                root.after(500, single_click, event, segments, sorted_pathlist[i_val][0])
                #break
            else:
                btns[i_val].config(bg="red2")
                root.after(500, single_click, event, segments, sorted_pathlist[i_val][0])
                #break
            if prev == 1:
                btns[i_val+1].config(bg="gray99", fg="purple3", font="Dosis")

    elif btnclicked == 2:
        if i_val >= len(sorted_pathlist)-1:
            i_val = len(sorted_pathlist)-1
            return
        if event:
            i_val += 1
            n2 = 0
            if sorted_pathlist[i_val] == 'deleted':
                for h in range(i_val+1, len(sorted_pathlist)):
                    if sorted_pathlist[h] != 'deleted':
                        n2 = h - i_val
                        break
                i_val += n2
                btns[i_val].config(bg="red2")
                btns[i_val-n2-1].config(bg="gray99", fg="purple3", font="Dosis")
                root.after(500, single_click, event, segments, sorted_pathlist[i_val][0])
            else:
                btns[i_val].config(bg="red2")
                root.after(500, single_click, event, segments, sorted_pathlist[i_val][0])
            if prev == 1:
                btns[i_val-1].config(bg="gray99", fg="purple3", font="Dosis")
    prev = 1


c=0
e=0
d=0
z=0
shown = 0
def make_counter():
    global shown
    if shown == 0:
        tkMessageBox.showinfo('What to do', 'Left-click to add a DC, Right-click to discard a DC.')
        shown = 1
    for widget in frm3.winfo_children():
        widget.destroy()
    datapts = []    #stores (x, y, color) tuple for every click 
    drawpts = []    #stores the dots
    undolist = []   #delets the dots
    undopts = []    #stores the deleted (x, y, color) tuple
    global c
    c=0
    global d
    d=0
    global e
    e=0
    global z
    z=0
    for i in range(0, len(scoredImages)):
        if scoredImages[i][0] == filename_ext:
            temp_i = i
            break

    def clickm(event):
        datapts.append((event.x, event.y, 'red'))
        x1 = event.x-5
        y1=event.y-5
        x2=event.x+5
        y2=event.y+5
        id=canvas.create_oval(x1,y1,x2,y2,fill='',width=3,outline='red')
        drawpts.append(id)
        f =open(pathss,'a')
        f.write('{}  {}  {}  {}  {} {}\n'.format(x1,y1,x2,y2,0,id))
        f.close
        global c  
        c=c+1
        counter2.set(mc_and_dc_list[temp_i][0] + c - d)
        #global e 
        #e=e+1
        counter3.set(mc_and_dc_list[temp_i][0]+mc_and_dc_list[temp_i][1]+c-d)



    def clickd(event):
        datapts.append((event.x, event.y, 'green'))
        x1 = event.x-1
        y1=event.y-1
        x2=event.x+1
        y2=event.y+1
        global d  
        d=d+1
        if (mc_and_dc_list[temp_i][0] + c - d < 0):
            return
        id=canvas.create_oval(x1,y1,x2,y2,fill='green',width=3,outline='green')
        drawpts.append(id)
        f =open(pathss,'a')
        f.write('{}  {}  {}  {}  {} {}\n'.format(x1,y1,x2,y2,0,id))
        f.close
        counter2.set(mc_and_dc_list[temp_i][0] + c - d)
        #global e 
        #e=e+1
        counter3.set(mc_and_dc_list[temp_i][0] + mc_and_dc_list[temp_i][1] + c - d)

    def undo():
        if len(drawpts) > 0:
            f =open(pathss,'r')
            f.seek(0, 0)
            data = np.genfromtxt(pathss)
            f.close()
            lastdraw = drawpts.pop()
            lastpt = datapts.pop()
            undopts.append(lastpt)
            undolist.append(canvas.delete(lastdraw))
            i=0
            
            if lastpt[2] == 'red':
                global c
                c -= 1
            elif lastpt[2] == 'green':
                global d
                d -= 1
            counter2.set(mc_and_dc_list[temp_i][0] + c - d)
            counter3.set(mc_and_dc_list[temp_i][0] + mc_and_dc_list[temp_i][1] + c - d)

    def redo():
        if len(undolist) > 0:
            lastpt = undopts.pop()
            datapts.append(lastpt)
            drawpts.append(canvas.create_oval(lastpt[0]-1, lastpt[1]-1, lastpt[0]+1, lastpt[1]+1,fill=lastpt[2],width=3,outline=lastpt[2]))
            undolist.pop()
            if lastpt[2] == 'red':
                global c
                c += 1
            elif lastpt[2] == 'green':
                global d
                d += 1
            counter2.set(mc_and_dc_list[temp_i][0] + c - d)
            counter3.set(mc_and_dc_list[temp_i][0] + mc_and_dc_list[temp_i][1] + c - d)

    def save():
        file_name = filename_ext + '_modified.jpg'
        savedialog = tkMessageBox.askokcancel("Save", "Are you sure you want to save the changes?")
        time.sleep(0.30)
        if savedialog is True:
            x00 = canvas.winfo_rootx()
            y00 = canvas.winfo_rooty()
            x11 = x00 + canvas.winfo_width()
            y11 = y00 + canvas.winfo_height()
            im = ImageGrab.grab((x00, y00, x11, y11))
            im.save(file_name)
            tkMessageBox.showinfo("Saved!", "File saved as " + file_name)
            print "saved"
        #canvas.update()
        #ps = canvas.postscript(file = file_name, colormode = 'color')
    

    canvas.create_image(0, 0, image = photo1, anchor = NW)   
    canvas.bind("<Button-1>",clickm)
    canvas.bind("<Button-3>",clickd)
    #canvas.bind("Key",clicka)
    
    counter1=IntVar()
    counter1.set(mc_and_dc_list[temp_i][1])
    #label0=Label(frm3,bg="white",text="Click to add a DC\nDouble-click to\ndiscard a DC",width=15)
    #label0.pack(fill=X, pady=10)
    label1=Label(frm3,bg="white",text="Monocentric",width=15)
    label1.pack(fill=X, pady=10)
    label2= Label(frm3,bg="#80c1ff",textvariable=counter1,width=15)
    label2.pack(fill=X, pady=10)

    counter2=IntVar()
    counter2.set(mc_and_dc_list[temp_i][0])
    label3=Label(frm3,bg="white",text="Dicentric",width=15)
    label3.pack(fill=X, pady=10)  
    label4= Label(frm3,bg="#80c1ff",textvariable=counter2,width=15)
    label4.pack(fill=X, pady=10)

    #savebtn = Tkinter.Button()
#    counter4=IntVar()
#    counter4.set(0)
#    label5=Label(frm3,bg="white",text="Acentric",width=15)
#    label5.pack(fill=X, pady=10)
#    label6= Label(frm3,bg="#80c1ff",textvariable=counter4,width=15)
#    label6.pack(fill=X, pady=10)

    counter3=IntVar()
    counter3.set(mc_and_dc_list[temp_i][0]+mc_and_dc_list[temp_i][1])
    label7=Label(frm3,bg="white",text="Total Segments",width=15)
    label7.pack(fill=X, pady=10)
    label8= Label(frm3,bg="#80c1ff",textvariable=counter3,width=15)
    label8.pack(fill=X, pady=10)

    def manual_file():
        soft_data=openpyxl.Workbook()
        soft_sheet = soft_data.active
        soft_sheet.title="manual_scoring"
        c1 = soft_sheet['A1'] 
        c1.value = "Metaphase_ID" 
        c2 = soft_sheet['B1'] 
        c2.value = "Monocentric"
        c3 = soft_sheet['C1'] 
        c3.value = "Dicentric" 
        c4 = soft_sheet['D1'] 
        c4.value = "Acentic"
        c5 = soft_sheet['E1'] 
        c5.value = "Total"
        c12 = soft_sheet['A2'] 
        c12.value = file_name1
        c22 = soft_sheet['B2'] 
        c22.value = c
        c32 = soft_sheet['C2'] 
        c32.value = d
        c42 = soft_sheet['D2'] 
        c42.value = z
        c52 = soft_sheet['E2'] 
        c52.value = e
        soft_data.save("manual_scoring%s.xlsx"%time.time())

    button10=Button(frm3,bg="white",text="Undo",width=15, command=undo)
    button10.pack(fill=X,pady=10)
    button11=Button(frm3,bg="white",text="Redo",width=15, command=redo)
    button11.pack(fill=X,pady=10)
    button9=Button(frm3,bg="white",text="Save",width=15,command=save)
    button9.pack(fill=X,pady=10)


def display_metaphase():
    images = []
    buttons = []
    row_number = 0
    iteration = 0
    for widget in scframe.interior.winfo_children():
        widget.destroy()
    canvas.delete("all")
    min_score=threshold_entry.get()
    print "min_score"
    print min_score

    sorted_pathlist=sorted(originalPath, key=lambda x:x[1][0])
    temp=0
    flag=0
    for k, v in sorted_pathlist:
        print k,v
        print v[0], min_score

        if int(v[0])<=int(min_score):
            print v[0], min_score
        # break
            for x in xrange(len(v)):
                v[x] = str(v[x])
            text = k
            segments = v[1]
            print text
            if temp != v[0]:
                iteration=0
                row_number+=1
#                setofimgs=set()
                texttodisp = 'Rank: ' + v[0]
                w=Tkinter.Label(scframe.interior, text=texttodisp, height=1, width=30, bg="gray", fg="black", relief=Tkinter.FLAT)
                w.grid(row= row_number, columnspan=2)
                row_number+=1
#                tkvar = StringVar(root) 
            image = Image.open(text)
            image = image.resize((100,100), Image.ANTIALIAS)
            photo = ImageTk.PhotoImage(image)
            btn = Tkinter.Button(scframe.interior,height=2, width=20, relief=Tkinter.FLAT, bg="gray99", fg="purple3",font="Dosis")
            btn.config(image=photo,width="100",height="100")
            btn.image=photo
            btn.grid(row=row_number, column=iteration)
            images.append(photo)
            buttons.append(btn)
            if iteration==1:
                row_number+=1
            iteration = 1-iteration
#            setofimgs.add(btn)
#            popupMenu = OptionMenu(root, tkvar, *setofimgs)
#            popupMenu.pack()
#            tkvar.trace('w', change_dropdown)
            btn.bind('<Button-1>', lambda event, segments = segments, text=text :test(event, sorted_pathlist,buttons,segments,text,0,1))
            btn.bind('<Double-Button-1>', lambda event, segments = segments, text=text :test(event,sorted_pathlist,buttons,segments,text,0,101))
            temp=v[0]
    prevBtn = Button(f3, text='<', bg='gray99', fg='black')
    prevBtn.place(relx=0.80, rely=0.99, anchor=SE)
    prevBtn.bind('<Button-1>', lambda event, segments = segments, text=text :test(event,sorted_pathlist,buttons,segments,text,1, 1))
    nextBtn = Button(f3, text='>', bg='gray99', fg='black')
    nextBtn.place(relx=0.90, rely=0.99, anchor=SE)
    nextBtn.bind('<Button-1>', lambda event, segments = segments, text=text :test(event,sorted_pathlist,buttons,segments,text, 2, 1))


def display_rank():
    plt.bar(range(1,len(rankPlot)+1),rankPlot)
    plt.title('Rank Statistics')
    plt.ylabel('Number of metaphases')
    plt.xlabel('Rank of metaphase')
    plt.grid(True,color='k')
    plt.show()
    plt.close()

def image_path():           # For Input

    global file_name_list,f2,f3, root, scframe, frm1,frm2,frm3,frm4,frm5,frm6, canvas, canvas2
    root, f2, f3, scframe, canvas, f4,  canvas2, frm1,frm2,frm3,frm4,frm5,frm6 = make_gui_before()
    root.mainloop()
    file_name_list = []

    if selected_file != "":
        file_name_list.append(selected_file)
    if selected_folder != "":
        for file in os.listdir(selected_folder):
            if file.endswith(extension):
                file_name_list.append(os.path.join(selected_folder, file))

    file_name_list.sort()
    if not file_name_list:
        print "Select appropriate folder"
        sys.exit()
    else :

        global progress
        progress= ttk.Progressbar(f2, orient = 'horizontal', maximum = len(file_name_list), variable=downloaded, mode = 'determinate')
        progress.pack(fill=BOTH)
        start = ttk.Button(f2,text='Run ranking algorithm',command= lambda: Threader())
        start.pack(fill=BOTH)
        root.mainloop()
        b2=Button(f3, text = "Show Rank Statistics", command=display_rank)
        b2.pack(fill=X, pady = 30)
        
        b11=Button(f3, text="Manual Detection ",command=make_counter,bg="gray", fg="black", state=ACTIVE,width = 25).pack(fill=X, pady=10)

        global threshold_entry
        threshold_entry = Entry(f3, bd =5)
        threshold_entry.delete(0,'end')
        threshold_entry.insert(1, 10)
        threshold_entry.pack(fill = X)
        b = Button(f3, text = "Show metaphases with rank upto:",command=display_metaphase)
        b.pack(fill=X)
        global image_name_entry
        Label(f3, text="Selected Image : ", bg="gray", fg="black", width = 25).pack(fill=X, pady=10)
        image_name_entry = Entry(f3, bd =5)
        image_name_entry.pack(fill = X)
        global total_segments
        Label(f3, text="Number of segments", bg="gray", fg="black", width = 25).pack(fill=X, pady=10)
        total_segments = Entry(f3, bd =5)
        total_segments.pack(fill = X)
        global metaphase_rank
        Label(f3, text="Rank", bg="gray", fg="black", width = 25).pack(fill=X, pady=10)
        metaphase_rank = Entry(f3, bd =5)
        metaphase_rank.pack(fill = X)
        global b4
        b4=Button(f3, text = "Detect DCs")
        b4.pack(fill=X, pady = 30)
        b4.bind('<Button-1>', lambda event, segments = '', text='' :detect_dc(event,segments,text,1))
        root.mainloop()

        return file_name_list, f2, root


if __name__ == '__main__':
    red_cross = Image.open("Red-Cross-PNG-File.png")
#    red_cross = red_cross.crop((500, 500, 1500, 1500))
#    red_cross = red_cross.resize((240,240), Image.ANTIALIAS)

    file_name_list,f2,root = image_path()
#    print "originalPath: ",originalPath
#    print "mc_and_dc_list: ",mc_and_dc_list
#    print "scoredImages: ", scoredImages
#    print "scoredImagesList: ", scoredImagesList'''