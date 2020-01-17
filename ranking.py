# File Number 1

# All imports
import time
import numpy as np          # for+ numerical processing 
import argparse             # parse command line arguments
import cv2                  # opencv bindings
from matplotlib import pyplot as plt
from scipy.misc import imsave
import random
import scipy.ndimage
# import mahotas as mh
import os
from scipy.ndimage import gaussian_filter1d
import ttk
import PIL.Image
from os.path import basename
import sys
import Tkinter, Tkconstants, tkFileDialog

from Tkinter import *
import shutil           # For removing directory
import progressbar      # for progressbar
from scipy.interpolate import spline

from PIL import Image, ImageTk
import tkMessageBox
import skimage
from skimage.io import imread
import openpyxl
from openpyxl import Workbook

global soft_data, soft_sheet



soft_data = openpyxl.load_workbook('SoftDataUpload.xlsx')
soft_sheet = soft_data['SampleSoftData']

#red_cross = Image.open("Red-Cross-PNG-File.png")
# red_cross = red_cross.crop((500, 500, 1500, 1500))
#red_cross = red_cross.resize((240,240), Image.ANTIALIAS)


    
def fileout():
    return selected_file

def apply_filter(path):     # Applying Gaussian Filter
    # print path
    try:
        dna = cv2.imread(path)
    except IOError:
        print('There was an error opening the file!')
        sys.exit()
    # dnaf = cv2.GaussianBlur((dna,(5,5),0))
    # dnaf = scipy.ndimage.gaussian_filter(dna, 0.7)
    # dnaf_uint8 = dnaf.astype('uint8')
    # T = mh.thresholding.otsu(dnaf_uint8)
    # filtered = dnaf_uint8 > T
    # path_ext = os.path.splitext(path)[0]
    # imsave(path + '_filtered_image.jpg',filtered)
    imsave(path + '_filtered_image.jpg',dna)
    return

def get_coordinates(actual_image, c, coordinates_list):
    # fo.write(str(box))
    # coordinates_list.append(str(box))
    coordinates_list.append(c)
    rect = cv2.minAreaRect(c)
    box = cv2.cv.BoxPoints(rect)
    box = np.int0(box)
    w = rect[1][0]
    h = rect[1][1]
    Xs = [i[0] for i in box]
    Ys = [i[1] for i in box]
    x1 = min(Xs)
    x2 = max(Xs)
    y1 = min(Ys)
    y2 = max(Ys)

    angle = rect[2]
    if w > h :
        angle -= 90
        temp = h
        h = w
        w = temp

    angle = (np.pi*angle)/180
    center = ((x1+x2)/2,(y1+y2)/2)
    w1 = int(round(w))
    h1 = int(round(h))
    return center, angle, w1, h1,x1,y1,x2,y2

def create_dir(newpath):
    if not os.path.exists(newpath):
        #print(newpath)
        #print("sdfh kjdsbfj djghdjg jg djfg jdfgjdf gfdhg dfu ghdfjghdfhgidfhgjfdgdfjgd gjdh")
        os.makedirs(newpath)
        print "New Directory by Rajan  Created : " + newpath    
    return
def delete_file(filepath):
    try:
        os.remove(filepath)
    except: pass
def get_area_ratio(black_pixels,total_pixels):
    return float(float(black_pixels)/float(total_pixels))

def get_rect_ratio(width, average_width):
    return float(float(width)/float(average_width))

def get_hi_ratio(black_pixels, average_width, height):
    h_i = float(float(black_pixels)/float(average_width))
    hi_ratio = float(float(h_i)/float(height))
    return hi_ratio

def get_max_ratio(width, new_average_width):
    return float(float(width)/float(new_average_width))
def crop_image(image, centre, theta, width, height):
    output_image = cv2.cv.CreateImage((width, height), image.depth, image.nChannels)
    mapping = np.array([[np.cos(theta), -np.sin(theta), centre[0]], [np.sin(theta), np.cos(theta), centre[1]]])
    map_matrix_cv = cv2.cv.fromarray(mapping)
    cv2.cv.GetQuadrangleSubPix(image, output_image, map_matrix_cv)
    # cv2.cv.WarpAffine(image, output_image, map_matrix_cv)
    return output_image

mc_and_dc_list = []
def make_segments(count,thresh,blur,image, path, newpath, actual_contours_path,segment_path):
    # print image.shape
    # fo = open("processing.txt", "a")
    # print path
    coordinates_of_segments={}
    temp_coordinates=[]
    ret,th = cv2.threshold(blur,thresh,255,cv2.THRESH_BINARY)

    # individual = 'croppedSegments/'
    xpath = newpath
    imsave(newpath + '/thresholded_image.jpg',th)
    filename = list(path.split('/'))
    filename=filename[-1]
    filename = list(filename.split('\\'))
    filename=filename[-1]
    # print filename
    imsave(thresholded_path+filename, th)

    cnts,hierarchy = cv2.findContours(th, 1, 2)

    thresholded_image = cv2.cv.LoadImage(newpath + '/thresholded_image.jpg')
    pathses=newpath +"/size.txt"
    # delete_file(newpath + '/thresholded_image.jpg')
    actual_image = cv2.imread(path)
    file_name =[]
    file_name = path.split("\\")
    file = file_name[-1]
    coordinates_of_segments[file] = []    
    # print file,thresh
    loaded_image = cv2.cv.LoadImage(path)
    rect_image = actual_image
    contour_list = []
    mask = np.ones(image.shape[:2], dtype="uint8") * 255
     
    # loop over the contours
    number =  0
    red_number = 0
    green_number = 0
    global mc_and_dc_list
    segment_list = []
    original_segment_list = []
    coordinates_list = []
    xmin=10000
    ymin=10000
    xmax=0
    ymax=0
    w1=0
    h1=0
    counters=0
    for c in cnts:
        
        approx = cv2.approxPolyDP(c,0.009*cv2.arcLength(c,True),True)
        area = cv2.contourArea(c)

        if ((len(approx) > 8) & (area < 4000) & (area > 100)):

            number += 1  
            global w,h  
            center, angle, w, h,x1,y1,x2,y2 = get_coordinates(actual_image, c, coordinates_list)
            if(x1-w <xmin):
                xmin=x1-w
            if(x2+w >xmax):
                xmax=x2+w
            if(y1-h <ymin):
                ymin=y1-h
            if(y2+h >ymax):
                ymax=y2+h
            w1+=w
            h1+=h
            counters=counters+1
#            print(x1-w,y1-h,x2+w,y2+h,counters,h1,h)
            crop_th = crop_image(thresholded_image, center, angle, w, h)
            crop = crop_image(loaded_image, center, angle, w, h)
            image = crop
            # create_dir(individual+file+'/')
            cv2.cv.SaveImage(newpath + '/' + 'contour_' + str(number) + '.jpg',crop_th)
            cv2.cv.SaveImage(actual_contours_path + 'contour_' + str(number) + '.jpg',crop)
            # cv2.cv.SaveImage(individual+ file+'/' + str(random.randint(1,50000))  + '.jpg',crop)
            temp_image = PIL.Image.open(newpath + '/' + 'contour_' + str(number) + '.jpg')
            original_temp_image = PIL.Image.open(actual_contours_path + 'contour_' + str(number) + '.jpg')
            segment_list.append(temp_image)
            original_segment_list.append(original_temp_image)
            
            # image = original_temp_image
            image = skimage.color.rgb2gray(skimage.io.imread(actual_contours_path + 'contour_' + str(number) + '.jpg'))
            delete_file(newpath + '/' + 'contour_' + str(number) + '.jpg')
            delete_file(actual_contours_path + 'contour_' + str(number) + '.jpg')
            total=[]
            h=image.shape[0]
            w=image.shape[1]

            for x in xrange(h):
                s=0
                for y in xrange(w):
                    s+=image[x][y]
                total.append(s)

            avg = [sum(total)/len(total)]*len(total)
            T = list(range(len(total)))
            t = np.array(T)
            power = np.array(total)
            totalnew = np.linspace(t.min(),t.max(),len(total))
            power_smooth = spline(t,power,totalnew)

            # ax = axs[1]
            sigma = 3
            x_g1d = gaussian_filter1d(totalnew, sigma)
            y_g1d = gaussian_filter1d(power_smooth, sigma)

            index = []
            temp=0
            for i in xrange(1,len(y_g1d)-1):
                if y_g1d[i]>y_g1d[i-1] and y_g1d[i]>y_g1d[i+1]:
                    index.append(i)
                    
            if len(index)==0:
                x_g1d=totalnew
                y_g1d=power_smooth
                for i in xrange(1,len(y_g1d)-1):
                    if y_g1d[i]>y_g1d[i-1] and y_g1d[i]>y_g1d[i+1]:
                        index.append(i)

            cm = []

            for x in xrange(1,len(index)):
                for y in xrange(x):
                    if y_g1d[index[y]]<y_g1d[index[x]]:
                        temp = index[y]
                        index[y] = index[x]
                        index[x] = temp


            if len(index)>0:
                mx = [y_g1d[index[0]]]*len(total)
            # plt.plot(t,mx)
                cent1 = index[0]
            # ax=axs[0]
                cm1 = (w/2,cent1)
                cv2.circle(image,cm1,3,(0,1,0),-1)
                cm.append(cm1)
            DCcount=0

            if len(index)>1 and y_g1d[index[1]]>avg[0] and abs(y_g1d[cent1]-y_g1d[index[1]])<abs(y_g1d[index[1]]-y_g1d[int(avg[0])]):
            # if len(index)>1 and total[index[1]]>avg[0] and abs(cent1-index[1])>h/a and abs(total[cent1]-total[index[1]])<abs(total[index[1]]-total[int(avg[0])]):
                mx2 = [y_g1d[index[1]]]*len(total)
                # plt.plot(t,mx2)
                cent2 = index[1]
                cm2 = (w/2,cent2)
                cv2.circle(image,cm2,3,(0,1,0),-1)
                cm.append(cm2)
                DCcount+=1

            if len(cm)==2:
                red_number += 1
                rect = cv2.minAreaRect(c)
                box = cv2.cv.BoxPoints(rect)
                temp_box=list(box)
                temp_box.append((1,1))
                box = np.int0(box)
                
                temp_coordinates.append(temp_box)
                # print "coordinates of dc"
                # print box
                cv2.drawContours(actual_image,[box],0,(0,0,255),2)
            else:
                green_number += 1
                rect = cv2.minAreaRect(c)
                box = cv2.cv.BoxPoints(rect)
                temp_box=list(box)
                temp_box.append((0,0))
                box = np.int0(box)

                temp_coordinates.append(temp_box)
                # print "coordinates of mc"
                # print box
                cv2.drawContours(actual_image,[box],0,(0,255,0),2)
    f=open(pathses,'w')
    ##print(w1,h1,counters,w1/(counters),h1/(counters))
    #print(pathses)
    #print("dsfj sdjfhjdsfdsfjdsjf jdsfjds jfdsjjf sjfdjsfjdsh")
    f.write('{}  {}  {}  {}  {}  {}'.format(xmin,ymin,xmax,ymax,w1/(counters),h1/(counters)))
    f.close()
    coordinates_of_segments[file] = temp_coordinates
    cv2.imwrite(segment_path+filename, actual_image)
    time.sleep(0.30)
    #cv2.waitKey(50)
    print "^^^^^^^^^^^^^^",time.time()
    i=count+1
    # while soft_sheet['A'+str(i)]==None:
    soft_sheet['A'+str(i)]=file
    soft_sheet['B'+str(i)]=number
    if not DCcount==0:
        soft_sheet['C'+str(i)]=DCcount
    pathses,tail=os.path.split(pathses)
    pathses,tail=os.path.split(pathses)
    soft_data.save('data.xlsx')
    mc_and_dc_list.append([red_number, green_number, number,tail[8:]])
    return mc_and_dc_list, number, segment_list, original_segment_list, coordinates_list, coordinates_of_segments

def sequence(count,path):               # Flow Sequence
    global dir  
    dir = os.path.dirname(path) 
    # print dir                            # present directory
    filename_ext = os.path.splitext(basename(path))[0]      # filename without extension
    # global rank1, rank2, rank3, rank4, rank5, rank6, rank7, rank8, rank9, rank10
    create_dir(dir +'/segments/score1/')
    create_dir(dir +'/segments/score2/')
    create_dir(dir +'/segments/score3/')
    create_dir(dir +'/segments/score4/')
    create_dir(dir +'/segments/score5/')
    create_dir(dir +'/segments/score6/')
    create_dir(dir +'/segments/score7/')
    create_dir(dir +'/segments/score8/')
    create_dir(dir +'/segments/score9/')
    create_dir(dir +'/segments/score10/')
    apply_filter(path)                                      # apply filter on the image
    image = cv2.imread(path + '_filtered_image.jpg',0)      # open filtered image and store it in variable "image"
    delete_file(path + '_filtered_image.jpg')               # delete filtered image from disk
    newpath = dir + "/results_" + filename_ext              # path for storing results
    print(newpath)
    create_dir(newpath)                                     # creating directory for results
    actual_contours_path = newpath + '/actual/'             # path where actual segments are saved
    create_dir(actual_contours_path)                        # creating directory for actual segments
    segment_path = dir + "/segments/"
    global thresholded_path
    thresholded_path = dir + "/binary/"
    # print dir
    # print thresholded_path
    create_dir(thresholded_path)
    mc = actual_contours_path+'/mc/'
    create_dir(mc)
    dc = actual_contours_path+'/dc/'
    create_dir(dc)  
    create_dir(segment_path)                        # path where bounded segments are saved
    thresh,blur = threshold(image)
    mc_and_dc_list, number, segment_list, original_segment_list, coordinates_list, coordinates_of_segments = make_segments(count,thresh,blur,image, path, newpath, actual_contours_path,segment_path)    # process image and get no. of segments in image and list of segments
    return mc_and_dc_list, number, segment_list, original_segment_list, coordinates_list, actual_contours_path, newpath, segment_path, coordinates_of_segments

def get_num_pixels(image): 
    width, height = image.size
    total_pixels = width*height
    black_pixels = 0
    white_pixels = 0
    pix = image.load()
    for y in xrange(height):
        for x in xrange(width):
            if pix[x,y] < (127,127,127):
                black_pixels = black_pixels + 1
            else:
                white_pixels = white_pixels + 1
    return total_pixels, black_pixels, white_pixels, height, width

def calculate_perimeters(number, segment_list):
    # max_width = 0
    count = 0.0000001   
    total_width = 0.0000001
    new_total_width = 0;
    Area_Ratio_List = []
    for i in xrange(number):
        total_pixels, black_pixels, white_pixels, height, width = get_num_pixels(segment_list[i])
        # print total_pixels, black_pixels, white_pixels
        new_total_width += width
        Area_Ratio = get_area_ratio(black_pixels, total_pixels)
        Area_Ratio_List.append(Area_Ratio)
        if Area_Ratio > 0.6784:  #0.59062:  #0.63
            total_width += width
            count += 1
    average_width = float(float(total_width)/float(count))
    new_average_width = float(float(new_total_width)/float(number))

    W_Rect_Ratio_List = []
    H_i_Ratio_List = []
    W_Max_Ratio_List = []
    for i in xrange(number):
        total_pixels, black_pixels, white_pixels, height, width = get_num_pixels(segment_list[i]) 
        Rect_Ratio = get_rect_ratio(width, average_width)
        W_Rect_Ratio_List.append(Rect_Ratio)
        hi_ratio = get_hi_ratio(black_pixels, new_average_width, height)
        H_i_Ratio_List.append(hi_ratio)
        max_ratio = get_max_ratio(width, new_average_width)
        W_Max_Ratio_List.append(max_ratio)
    return  Area_Ratio_List, W_Rect_Ratio_List, H_i_Ratio_List, W_Max_Ratio_List
#combined.py
def classify_images(number, Area_Ratio_List, W_Rect_Ratio_List, H_i_Ratio_List, W_Max_Ratio_List, original_segment_list, coordinates_list):
    class1 = []
    class2 = []
    class3 = []
    class4 = []
    class1_coordinates = []
    class2_coordinates = []
    for i in xrange(number):
        if Area_Ratio_List[i] >= 0.63: #0.63968441:     #0.63

            if W_Rect_Ratio_List[i] < 0.88:     #0.88
                class4.append(original_segment_list[i])
            elif W_Rect_Ratio_List[i] > 1.4: #1.4:
                class4.append(original_segment_list[i])
            else :
                class1.append(original_segment_list[i])
                class1_coordinates.append(coordinates_list[i])
        else :
            if H_i_Ratio_List[i] < 0.6:     #0.6
                class4.append(original_segment_list[i])
            elif W_Max_Ratio_List[i] > 1.38603:     #1.5
                class3.append(original_segment_list[i])
            else :
                class2.append(original_segment_list[i])
                class2_coordinates.append(coordinates_list[i])
    return class1, class2, class3, class4, class1_coordinates, class2_coordinates

def save_segments_in_classes(actual_contours_path, class1, class2, class3, class4):

    for i in xrange(len(class1)):
        imsave(actual_contours_path +'/'+ str(i+1) + '.jpg',class1[i])
    for i in xrange(len(class2)):
        imsave(actual_contours_path +'/'+ str(len(class1)+i+1) + '.jpg',class2[i])
    for i in xrange(len(class3)):
        imsave(actual_contours_path +'/'+ str(len(class1)+len(class2)+i+1) + '.jpg',class3[i])
    for i in xrange(len(class4)):
        imsave(actual_contours_path +'/'+ str(len(class1)+len(class2)+len(class3)+i+1) + '.jpg',class4[i])
    return
files = iter(np.arange(1,10000))


def find_good_metaphases(downloaded, progress,file_name_list,f2,root,threshold_value):
    threshold = int(threshold_value)
    good_metaphases = 0   
    files = iter(np.arange(1,len(file_name_list)+1))                 
    bar = progressbar.ProgressBar(maxval=len(file_name_list), widgets=[progressbar.Bar('=', '[', ']'), ' ', progressbar.Percentage()])
    bar.start()
    ans_list = []       # list for showing output in terminal
    good_metaphases_list = {} # dictionary, would be used in making gui
    good_metaphases_list_with_coordinates = {}
    ans = ""
    count = 1
    segpath = []
    dict_of_coordinates={}
    for i in xrange(len(file_name_list)):
        try:
            downloaded.set(next(files)) # update the progress bar
            # root.after(1, loading) # call this function again in 1 millisecond
        except StopIteration:
            print "100% progress"
         
        path = file_name_list[i]
        filename_ext = os.path.splitext(basename(path))[0]      # filename without extension
        mc_and_dc_list, number, segment_list, original_segment_list, coordinates_list, actual_contours_path, newpath,segment_path, coordinates_of_segments = sequence(count,path)
        dict_of_coordinates[filename_ext+'.JPG'] = coordinates_of_segments[path]
        # print "coordinates list"
        # print coordinates_list
        segpath.append(actual_contours_path)
        count+=1
        if number > 0 and number < 100:
            Area_Ratio_List, W_Rect_Ratio_List, H_i_Ratio_List, W_Max_Ratio_List = calculate_perimeters(number, segment_list)
            class1, class2, class3, class4, class1_coordinates, class2_coordinates = classify_images(number, Area_Ratio_List, W_Rect_Ratio_List, H_i_Ratio_List, W_Max_Ratio_List, original_segment_list, coordinates_list)
            # save_segments_in_classes(actual_contours_path, class1, class2, class3, class4)
            good_metaphases += 1
            ans = str(len(class1)) + " " + str(len(class2)) + " " + str(len(class3)) + " " + str(len(class4))
            good_metaphases_list[file_name_list[i]] = ans
            ans = filename_ext + " : " + ans                
            class_1_and_2_contours = class1_coordinates + class2_coordinates
            good_metaphases_list_with_coordinates[file_name_list[i]] = class_1_and_2_contours

        bar.update(i+1)
    bar.finish()
    #print(mc_and_dc_list)
    #print("dskf d hfdsfdsf hdsfdsf dsufhsdhf kdhfjdshfdhf hdsfh dsgfhjdhjfdshfdsfdsf dsgfdfgdhjfgdsg fjdshfjdhjfdshj")
    return mc_and_dc_list, number, good_metaphases_list, good_metaphases_list_with_coordinates, segpath, dict_of_coordinates


def threshold(image):
    # print image
    blur = cv2.GaussianBlur(image,(5,5),0)
    # print blur
    hist = cv2.calcHist([blur],[0],None,[180],[0,180])
    hist_norm = hist.ravel()/hist.max()
    Q = hist_norm.cumsum()

    bins = np.arange(180)

    fn_min = np.inf
    thresh = -1

    for i in xrange(1,180):
        p1,p2 = np.hsplit(hist_norm,[i]) # probabilities
        q1,q2 = Q[i],Q[179]-Q[i] # cum sum of classes
        b1,b2 = np.hsplit(bins,[i]) # weights

        # finding means and variances
        m1,m2 = np.sum(p1*b1)/q1, np.sum(p2*b2)/q2
        v1,v2 = np.sum(((b1-m1)**2)*p1)/q1,np.sum(((b2-m2)**2)*p2)/q2

        # calculates the minimization function
        fn = v1*q1 + v2*q2
        # print 'fn',fn
        if fn < fn_min:
            fn_min = fn
            thresh = i

    return thresh,blur

def rank(folder):
    # print folder
    # folder = 'seg3/'
    global soft_data, soft_sheet
    print(os.getcwd())
    
    scoredImages = []
    originalPath = []
    rankPlot = [0]*10 
    for path in folder:
        paths=path
        filename1 = path.split('results_')
        filename = filename1[-1].split('./')
        temp = filename1[0]+filename[0]+'..JPG'
        # print temp
        width = []
        files = []
        # print path
        count = 0
        for file in os.listdir(path):
            if file.endswith('.jpg'):
                count+=1
                files.append(file)
                # print file
                img = cv2.imread(path+file)
                width.append(img.shape[1])
        if count==46:
            score = 10
        elif 44<=count<=48:
            score = 9
        elif 42<=count<=50:
            score = 8
        elif 41<=count<=51:
            score = 7
        elif 40<=count<=52:
            score = 6
        else:
            score = 5
        
        med = np.median(width)
        overlap_count = 0
        for file in files:
            img = cv2.imread(path+file)
            # print file
            x = img.shape[1]
            if x>med+10:
                overlap_count+=1
            #   os.remove(path+file)
        
        c=0
        for file in os.listdir(path):
            #c=0
        #   print file
            if file.endswith('.jpg'):
            #if file.endswith('.jpg'):
                img = cv2.imread(path+file)
                img_g = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                h = img_g.shape[0]
                w = img_g.shape[1]
                th2 = cv2.adaptiveThreshold(img_g,255,cv2.ADAPTIVE_THRESH_MEAN_C,cv2.THRESH_BINARY,15,0)
                total = []
                for x in xrange(w):
                    sum=0
                    for y in xrange(h):
                        sum+=th2[y][x]
                    total.append(sum)
                avg = [np.sum(total)/len(total)]*len(total)
                val=avg[0]
                counter=0
                for p in range(len(total)-1):
                    if (val<total[p] and val>total[p+1]) or (val>total[p] and val<total[p+1]):
                        counter+=1
                        #os.remove(path+file)
                # if counter==2:
                #   c+=1
                #   os.remove(path+file)
        print c
                # print file,"not a single straight chromosome"
        score-=overlap_count
        score = max(1,score)
        if (c<.5*count):
            score=score
        elif(.5*count<c<.6*count):
            score-=1
        elif(.6*count<c<.7*count):
            score-=2
        elif(.7*count<c<.8*count):
            score-=3
        elif(.8*count<c<.9*count):
            score-=4
        elif(.9*count<c<count):
            score-=5
        score = max(1,score)            
        rankPlot[11-score-1]+=1 
        scoredImages.append((filename[-2]+'.', [11-score,count,overlap_count]))
        originalPath.append((temp, [11-score, count])) 

#    for u in range(0, len(originalPath)):
#        #img1 = cv2.imread(temp)
#        print "**********",originalPath[u][0]
#        q = originalPath[u][0].split('/')
#        imgname = q[len(q) - 1]
#        segmented_img = ''
#        for w in range(0, len(q)-1):
#            segmented_img = segmented_img + '/' + str(q[w])
#        segmented_img = segmented_img + '/segments/' + imgname
 #       img1 = cv2.imread(segmented_img)
 #       print segmented_img
#        #print filename[-2]
 #       path_to_store = dir + '/segments/score' + str(originalPath[u][1][0]) + '/' + str(imgname)
  #      print path_to_store
   #     status = cv2.imwrite(path_to_store, img1)
    #    time.sleep(0.1)

    i=1
    tps = 0
    scoredImagesList = []
    head,tail=os.path.split(paths)
    head,tail=os.path.split(head)
    head,tail=os.path.split(head)
    head=head+'/segments/scoredData.xlsx'
    if os.path.isfile(head):
        pass
    else:
        soft_data = openpyxl.load_workbook('Ranked.xlsx')
        soft_sheet = soft_data.active
        soft_data.save(head)
    soft_data = openpyxl.load_workbook(head)
    soft_sheet = soft_data.active
    
    for key, value in sorted(scoredImages, key=lambda (k,v): (v,k),reverse = True):
        if value[0]>5:
            tps+=value[1]
        os.chdir(folder[0])
        os.chdir("..")
        os.chdir("..")
        os.chdir("segments/")
        # print os.getcwd()
        # shutil.copy(key+".JPG", "score"+str(value[0]))
        scoredImagesList.append(key)
        c1=soft_sheet.cell(1+i,1)
        c1.value=key
        c1=soft_sheet.cell(1+i,2)
        c1.value=value[1]
        c1=soft_sheet.cell(1+i,3)
        c1.value=value[2]
        c1=soft_sheet.cell(1+i,4)
        c1.value=value[0]
       
        print "%s: %s" % (key, value)
        i+=1
    print tps
    pres_time = time.time()
    # print scoredImages.keys()
    #c1=soft_sheet.cell(1+1,8)
    #c1.value="dsjhfjkdshfjd"
    #print("sdk fdfdfdfhdshf dgshfghdjsgfhjdsghfgdshjfg hdg fhjsdghfgdshgfhdsgfdgshjfgdhs gfhdjsgfh")
    soft_data.save(head)
    #print(rankPlot);
    
    return originalPath,scoredImages, pres_time, scoredImagesList, rankPlot
