import Tkinter, Tkconstants, tkFileDialog
from Tkinter import *
from ranking import rank, find_good_metaphases
import os
import matplotlib.pyplot as plt
from PIL import Image, ImageTk, ImageGrab
from os.path import basename
import tkMessageBox
import shutil
import ttk
import threading
import numpy as np
import openpyxl
import time
import cv2
from openpyxl import Workbook

#import io
selected_file = ""
selected_folder = ""
undo_path=""
redo_path=""
mc_and_dc_list=""
dirtext='Select a Directory'
filetext= 'Select a file'

initial_directory = ""





class Metaphase_Analysis(Tkinter.Frame):
    def __init__(self, root):
        Tkinter.Frame.__init__(self, root)

        button_opt = {'fill': Tkconstants.BOTH, 'padx': 5, 'pady': 5}
        # self.file_button = Tkinter.Button(self, text = filetext, command = self.openfile, height = 6, width = 20)
        # self.file_button.pack(side=LEFT, **button_opt)
        self.dir_button = Tkinter.Button(self, text = dirtext,font=(None, 12), command = self.opendirectory, height = 5, width = 50)
        self.dir_button.pack(side=LEFT, **button_opt)
        # self.refine_button = Tkinter.Button(self, text = "Refine", command = self.refine)
        # self.refine_button.pack(side=LEFT, **button_opt)
        self.exit_button = Tkinter.Button(self, text = "Exit",font=(None, 12), command = root.destroy, height = 5, width = 20)
        self.exit_button.pack(side=LEFT, **button_opt)
        self.help_button = Tkinter.Button(self, text = "Help",font=(None, 12), command = self.helptext, height = 5, width = 20)
        self.help_button.pack(side=LEFT, **button_opt)
        # self.chr_label = Tkinter.Label(self,text="Chromosome segments", height = 6, width=25)
        # self.chr_label.pack(side = RIGHT, **button_opt)
        self.metaphase_label = Tkinter.Label(self,text="Metaphases", font=(None, 12), height = 5, width=25)
        self.metaphase_label.pack(side = LEFT, padx = 40)
        #self.pack(fill=BOTH, expand=NO)
        self.switch_button = Tkinter.Button(self, text = "Switch",font=(None, 12), command = self.helptext, height = 5, width = 25)
        self.switch_button.pack(side=LEFT, **button_opt)

        # define options for opening or saving a file
        self.file_opt = options = {}
        global extension
        extension = '.JPG'
        options['defaultextension'] = extension
        options['filetypes'] = [('image files', extension)]
        options['initialdir'] = initial_directory
        options['initialfile'] = ''
        options['parent'] = root
        options['title'] = 'Metaphase Analysis'

        # defining options for opening a directory
        self.dir_opt = options = {}
        options['initialdir'] = initial_directory
        options['mustexist'] = False
        options['parent'] = root
        options['title'] = 'MetaPhase Analysis'

    def helptext(self):
        #textmsg = '1. text to display\n2. text to display\n3. text to display\n4. text to display\n5. text to display'
        f1 = open("Help.txt", "r")
        textmsg = f1.read()
        showdialog = tkMessageBox.showinfo('Help', textmsg)
        f1.close()

    def openfile(self):
        filename = tkFileDialog.askopenfilename(**self.file_opt)
        self.file_button["text"] = str(filename) if filename else filetext
        if filename:
            global selected_file
            selected_file = str(filename)
        if self.file_button["text"] == str(filename):
            self.quit()

    def opendirectory(self):
        dirname = tkFileDialog.askdirectory(**self.dir_opt)
        self.dir_button["text"] = str(dirname) if dirname else dirtext
        if dirname:
            global selected_folder
            selected_folder = str(dirname)
        if self.dir_button["text"] == str(dirname):
            self.quit()


# Adding scrollbar to the canvas
class VerticalScrolledFrame(Tkinter.Frame):
    def __init__(self, parent, *args, **kw):
        Tkinter.Frame.__init__(self, parent, *args, **kw)            

        vscrollbar = Tkinter.Scrollbar(self, orient=Tkinter.VERTICAL)
        vscrollbar.pack(fill=Tkinter.Y, side=Tkinter.LEFT, expand=Tkinter.FALSE)
        canvas = Tkinter.Canvas(self, bd=0, highlightthickness=0, yscrollcommand=vscrollbar.set)
        canvas.pack(side=Tkinter.LEFT, fill=Tkinter.BOTH, expand=Tkinter.TRUE)
        vscrollbar.config(command=canvas.yview)

        # reset the view
        canvas.xview_moveto(0)
        canvas.yview_moveto(0)
        canvas.delete("All")
        # create a frame inside the canvas which will be scrolled with it
        self.interior = interior = Tkinter.Frame(canvas)
        interior_id = canvas.create_window(0, 0, window=interior, anchor=Tkinter.NW)
        self.pack(fill=BOTH, expand=YES)


        def myfunction(event):
            canvas.configure(scrollregion=canvas.bbox("all"),width=1240,height=400 )

        interior.bind('<Configure>', myfunction)

        def _configure_canvas(event):
            if interior.winfo_reqwidth() != canvas.winfo_width():
                # update the inner frame's width to fill the canvas
                canvas.itemconfigure(interior_id, width=canvas.winfo_width())
        canvas.bind('<Configure>', _configure_canvas)


def make_gui_before():
    root = Tkinter.Tk()
    global downloaded
    downloaded = IntVar()
    root.attributes('-fullscreen', True)

    root.title("Detection of dicentric chromosomes")
    w, h = root.winfo_screenwidth(), root.winfo_screenheight()
    posx  = 0
    posy  = 0
    root.wm_geometry("%dx%d+%d+%d" % (w, h, posx, posy))

    statusbar = Tkinter.Label(root, borderwidth=1, relief="sunken")
    statusbar.pack(side="bottom", fill="x")
    
    Metaphase_Analysis(root).pack()
    f2 = Frame(root, bg = "black", height=h/1.5, width = w/2)
    f2.pack(side=LEFT, fill = Y)
    f3 = Frame(root, bg = "black", height=h/1.5, width = w/9)
    f3.pack(side=LEFT, fill = Y)
    f4 = Frame(root, bg = "black", height=0, width = 0)
    f4.pack(side=RIGHT, fill = Y)


    frm1 = Frame(root,bg = "black", height = h/1.3,width = w/50)
    frm1.pack(side = RIGHT, fill=Y)
    frm2 = Frame(root,bg = "black", height = h/1.3,width = w/50)
    frm2.pack(side = RIGHT, fill=Y)

    frm3 = Frame(root,bg = "black", height = h/1.3,width = w/20)
    frm3.pack(side = RIGHT, fill=Y)
    frm4 = Frame(root,bg = "black", height = h/1.3,width = w/50)
    frm4.pack(side = RIGHT, fill=Y)
    frm5 = Frame(root,bg = "black", height = h/1.3,width = w/50)
    frm5.pack(side = RIGHT, fill=Y)
    frm6 = Frame(root,bg = "black", height = h/1.3,width = w/50)
    frm6.pack(side = RIGHT, fill=Y)

    scframe = VerticalScrolledFrame(root)
    scframe.pack()

    canvas = Canvas(f2, width = w/2, height = h/2)
    canvas.pack(fill=X, expand=2)
    canvas2 = Canvas(f4, width = 0, height = 0)
    canvas2.pack(fill=X, expand=2)

    return root, f2, f3, scframe, canvas ,f4,  canvas2, frm1,frm2,frm3,frm4,frm5,frm6


counter_red=0
counter_green=0
counter_total=0
fcounter_red=0
fcounter_green=0
fcounter_total=0

mz=0
dz=0
tz=0

class Threader(threading.Thread):
    
    def __init__(self, *args, **kwargs):
        
        threading.Thread.__init__(self, *args, **kwargs)
        self.daemon = True
        self.start()
    def run(self):
        root.quit()
        global originalPath,scoredImages,scoredImagesList, pres_time, rankPlot ,f3,counter_red,counter_green,counter_total,m
        global mc_and_dc_list, number_of_segments, good_metaphases_list, good_metaphases_list_with_coordinates, file_with_good_metaphases, actual_contours_path, dict_of_coordinates
        mc_and_dc_list, number_of_segments,good_metaphases_list, good_metaphases_list_with_coordinates,segpath, dict_of_coordinates = find_good_metaphases(downloaded, progress,file_name_list,f2,root,25)
        red=0
        green=0
        total=0
        originalPath,scoredImages, pres_time, scoredImagesList, rankPlot = rank(segpath)
        time.sleep(0.3)
        segment_path=''

        for i in range(0, len(originalPath)):
            #originalPath[i][0]
            rnk = originalPath[i][1][0]
            q = originalPath[i][0].split('/')
            imgname = q[len(q) - 1]
            segmented_img = ''
            for j in range(0, len(q)-1):
                segmented_img = segmented_img + '/' + str(q[j])
            segment_path=segmented_img+'/segments/scoredData.xlsx'
            segmented_img = segmented_img + '/segments/' + imgname
            img1 = cv2.imread(segmented_img)
            #print filename[-2]
            path_to_store = selected_folder + '/segments/score' + str(rnk)+ '/' + str(imgname)
            status = cv2.imwrite(path_to_store, img1)
        segment_path=segment_path[1:]
        wb = openpyxl.load_workbook(segment_path)
        ws = wb.active
        col=ws.max_column
        c1=ws.cell(2, 7)
        if c1.value is None:
            c1=ws.cell(1, col)
            c1.value="Dicentric"
            c1=ws.cell(1, col+1)
            c1.value="Monocentric"
            c1=ws.cell(1, col+2)
            c1.value="Total Chromsome"
            for k in range(0,ws.max_row-1):
                for i in range(2, ws.max_row+1 ):
                    c1=ws.cell(i,1)
                    if(mc_and_dc_list[k][3]==c1.value):
                        for j in range(col,col+4):
                            c1=ws.cell(i, j)
                            c1.value=mc_and_dc_list[k][j-col]
                       # else:
                        #    mc_and_dc_list[k][j-col]=c1.value
                        break
            wb.save(segment_path)
        row=ws.max_row
        col=ws.max_column
        for i in range(2, row+1 ):
            for j in range(col-3,col+1):
                c1=ws.cell(i, j)
                # print c1.value
                # print mc_and_dc_list[i-2][j-col+3]
                mc_and_dc_list[i-2][j-col+3]=c1.value
        for i in mc_and_dc_list:
            red+=i[0]
            green+=i[1]
            total+=i[2]
        
        counter_red=IntVar()
        counter_green=IntVar()
        counter_total=IntVar()
        counter_red.set(red)
        counter_green.set(green)
        counter_total.set(total)
        
        
        
        


class disp_segment:
    
    def __init__(self,root,image,i):
        if 0<=i<15:
            self.c = Canvas(frm1, height=60, width = 60)
            self.c.pack(fill=Y,expand=0)
            self.c.pack(fill=Y,expand=0)
            self.c.create_image(0,0,anchor = NW, image = image)
            self.c.image = image
        if 15<=i<30:
            self.c = Canvas(frm2, height=60, width = 60)
            self.c.pack(fill=Y,expand=0)
            self.c.pack(fill=Y,expand=0)
            self.c.create_image(0,0,anchor = NW, image = image)
            self.c.image = image
        if 30<=i<45:
            self.c = Canvas(frm3, height=60, width = 60)
            self.c.pack(fill=Y,expand=0)
            self.c.create_image(0,0,anchor = NW, image = image)
            self.c.image = image
        if 45<=i<60:
            self.c = Canvas(frm4, height=60, width = 60)
            self.c.pack(fill=Y,expand=0)
            self.c.pack(fill=Y,expand=0)
            self.c.create_image(0,0,anchor = NW, image = image)
            self.c.image = image
        if 60<=i<75:
            self.c = Canvas(frm5, height=60, width = 60)
            self.c.pack(fill=Y,expand=0)
            self.c.pack(fill=Y,expand=0)
            self.c.create_image(0,0,anchor = NW, image = image)
            self.c.image = image
        if 75<=i<90:
            self.c = Canvas(frm6, height=60, width = 60)
            self.c.pack(fill=Y,expand=0)
            self.c.pack(fill=Y,expand=0)
            self.c.create_image(0,0,anchor = NW, image = image)
            self.c.image = image
def show_segments(histPath):
    photos = []
    destroy(root)
    i=0
    for file in os.listdir(histPath):
        if file.endswith('.jpg'):
            img = ImageTk.PhotoImage(Image.open(histPath+file))  # PIL solution

            i+=1
def destroy(root):
    for widget in frm1.winfo_children():
        widget.destroy()
    for widget in frm2.winfo_children():
        widget.destroy()
    for widget in frm3.winfo_children():
        widget.destroy()
    for widget in frm4.winfo_children():
        widget.destroy()
    for widget in frm5.winfo_children():
        widget.destroy()
    for widget in frm6.winfo_children():
        widget.destroy()
def delete_dialogue_box(text):
    msg = "Are you sure to delete " + os.path.basename(text) + " ?"
    result = tkMessageBox.askquestion("Delete", msg, icon='warning')
    return result
def detect_dc(event,segments,text,extra=None):
    global single
    if extra == 1:
        single = True

        click('show dc', event, None,segments, text)
    make_counter()
    return
def removekey(d, key):
    r = dict(d)
    if key in r :
        del r[key]
    return r

filename_ext = 'a'
prev=0
w1=0
h1=0
def click(msg, event, _btn, segments, text):
    
    global path, segment_path,w1,h1
    if msg == 'single click':
        dir = os.path.dirname(text)
        global dc_path 
        segment_path = dir+'/segments/'
        dc_path = segment_path +os.path.basename(text)
        global file_name1
        file_name1=os.path.basename(text)
        global histPath, filename_ext
        filename_ext = os.path.splitext(basename(text))[0]
        histPath = dir + "/results_" + filename_ext + '/actual/'
        show_segments(histPath)
        canvas.delete("all")
        path = text
        canvas.pack(fill=Tkinter.BOTH, expand=2) # expand
        image = Image.open(dir+"/"+os.path.basename(text))       
        width, height = image.size
        #print(path)
        head,tail=os.path.split(undo_path)
        head=head+'/size.txt'
        
        data = np.genfromtxt(head)
        xmin,ymin,xmax,ymax,w1,h1=data
        #print(xmin,ymin,xmax,ymax)
        topleft = (xmin,ymin)
        bottomright = (xmax,ymax)
        cropped = image.crop((topleft[0], topleft[1], bottomright[0], bottomright[1]))
        width = bottomright[0] - topleft[0]
        height = bottomright[1] - topleft[1]
        image = cropped.resize((int(770), int(610)), Image.ANTIALIAS)
        #global photo1
        photo = ImageTk.PhotoImage(image)
        canvas.create_image(0, 0, image = photo, anchor = NW)
        canvas.photo = photo

        image_name_entry.delete(0, 'end')
        image_name_entry.insert(0, filename_ext)
        total_segments.delete(0, 'end')
        for i in range(0, len(scoredImages)):
            if scoredImages[i][0] == filename_ext:
                temp_i = i
                break
        total_segments.insert(0, str(scoredImages[temp_i][1][1]))
        metaphase_rank.delete(0, 'end')
        #print scoredImages[filename_ext][0]
        metaphase_rank.insert(0,str(scoredImages[temp_i][1][0]))
        
    elif msg == 'double click':
        images=[]
        #buttons=[]
        result = delete_dialogue_box(text)
        if result == 'yes':
            global prev
            prev = 2
            image = Image.open(text)
            image.paste(red_cross, (100, 100), red_cross)
            image = image.resize((100,100), Image.ANTIALIAS)
            photo = ImageTk.PhotoImage(image)
            event.widget.config(image=photo,width="100",height="100")
            images.append(photo)
            global good_metaphases_list, good_metaphases_list_with_coordinates
            good_metaphases_list = removekey(good_metaphases_list, text)
            good_metaphases_list_with_coordinates = removekey(good_metaphases_list_with_coordinates, text) 
            dir = os.path.dirname(text)
            #filename_ext = os.path.splitext(basename(text))[0]
            #print "filename_ext", filename_ext
            newpath = dir + "/results_" + filename_ext
            shutil.rmtree(newpath)

    elif msg == 'show dc':
        
        dir = os.path.dirname(text)
        path = segment_path+"/"+os.path.basename(text)
        canvas.delete("all")

        canvas.pack(fill=Tkinter.BOTH, expand=2) # expand
        image = Image.open(dc_path)

        path = text
        #filename_ext = os.path.splitext(basename(text))[0]
        histPath = dir + "/results_" + filename_ext + '/actual/'
        width, height = image.size
        head,tail=os.path.split(undo_path)
        head=head+'/size.txt'
        
        data = np.genfromtxt(head)
        xmin,ymin,xmax,ymax,w1,h1=data
        #print(xmin,ymin,xmax,ymax)

        topleft = (xmin,ymin)
        bottomright = (xmax,ymax)
        cropped = image.crop((topleft[0], topleft[1], bottomright[0], bottomright[1]))
        width = bottomright[0] - topleft[0]
        height = bottomright[1] - topleft[1]
        image = cropped.resize((int(680), int(600)), Image.ANTIALIAS)
        
        global photo1
        photo1 = ImageTk.PhotoImage(image)
        canvas.create_image(0, 0, image = photo1, anchor = NW)
        canvas.photo = photo1

    try:
        return histPath
    except Exception as NameError:
        print "nothing to return"


def test(event,sorted_pathlist,buttons, segments, text, btnclicked, extra=None):
    global single
    
        
    if extra == 1 and btnclicked == 0:
        single = True
        root.after(500, single_click, event, segments, text)
        
        highlightimg(event, sorted_pathlist,buttons, segments, text, btnclicked)
        
    if extra == 1 and (btnclicked == 1 or btnclicked == 2):
        single = True
        #root.after(500, single_click, event, segments, textpre)
        highlightimg(event, sorted_pathlist,buttons, segments, text, btnclicked)
        
    elif extra == 101:
        single = False
        click('double click', event,sorted_pathlist, segments, text)


def single_click(event, segments, text):
    global single
    if single:
        single = False

        click('single click', event, None,segments, text)


#i_val=-1
def doNothing(event):
    pass
def highlightimg(event, sorted_pathlist, btns, segments, text, btnclicked):
    global prev, i_val,undo_path,redo_path
    
    if btnclicked == 0:
        for i in range(0, len(sorted_pathlist)):
            if sorted_pathlist[i][0] == text:
                canvas.bind("<Button-1>",doNothing)
                canvas.bind("<Button-3>",doNothing)
                undo_path=text[:-4]
                head,tail=os.path.split(undo_path)
                undo_path=head +'/results_'+tail+'/undo.txt'
                redo_path=head +'/results_'+tail+'/redo.txt'
                f =open(undo_path,'a')
                data = np.genfromtxt(undo_path)
                
                if((0L,)==np.shape(data)):
                    f.write("x1  y1  x2  y2  colour  id  type\n")                    
                else:
                    pass
                f.close
                f =open(redo_path,'w')
                f.write("x1  y1  x2  y2  colour  id  type\n")
                f.close

                New_path = head + '/segments/color.xlsx'
                yyy =str(head)
                x = yyy.split("/")
                one_morepath = "/"
                for g in range(len(x)-2):
                    one_morepath = one_morepath + x[g+1] + "/"

                if os.path.isfile(New_path):
                    pass
                else:
                    soft_dat = openpyxl.Workbook()
                    soft_she = soft_dat.active
                    soft_dat.save(New_path)
                soft_dat = openpyxl.load_workbook(New_path)
                soft_she = soft_dat.active
                if event and prev != 2:
                    c1 = soft_she.cell(soft_she.max_row + 1, 1) 
                    c1.value = i
                    soft_dat.save(New_path)
                    btns[i].config(bg="red")
                    btns[i](background="red", foreground="red",activebackground="red", activeforeground="red")
                    if prev == 1:
                        btns[i_val].config(bg="gray99", fg="purple3", font="Dosis")
                    i_val = i
                if event and prev == 2:
                    c1 = soft_she.cell(soft_she.max_row + 1,1) 
                    c1.value = i
                    soft_dat.save(New_path)
                    btns[i].config(bg="red")
                    btns[i](background="red", foreground="red",activebackground="red", activeforeground="red")
                    i_val = i
                if i_val == i:
                    c1 = soft_she.cell(soft_she.max_row + 1, 1) 
                    c1.value = i
                    soft_dat.save(New_path)
                    btns[i].config(bg="red")
                    btns[i](background="red", foreground="red",activebackground="red", activeforeground="red")
                    i_val = i
                break
                soft_dat.save(New_path)

    elif btnclicked == 1:
        if i_val <= 0:
            i_val = 0
            return
        if event:
            i_val -= 1
            n1 = 0
            if sorted_pathlist[i_val] == 'deleted':
                for h in range(1, i_val+1):
                    if sorted_pathlist[i_val-h] != 'deleted':
                        n1 = h
                        break
                i_val -= n1
                btns[i_val].config(bg="red")
                btns[i_val](background="red", foreground="red",activebackground="red", activeforeground="red")
                btns[i_val+n1+1].config(bg="gray99", fg="purple3", font="Dosis")
                root.after(500, single_click, event, segments, sorted_pathlist[i_val][0])
                #break
            else:
                btns[i_val].config(bg="red")
                btns[i_val](background="red", foreground="red",activebackground="red", activeforeground="red")
                root.after(500, single_click, event, segments, sorted_pathlist[i_val][0])
                #break
            if prev == 1:
                btns[i_val+1].config(bg="gray99", fg="purple3", font="Dosis")
        

    elif btnclicked == 2:
        if i_val >= len(sorted_pathlist)-1:
            i_val = len(sorted_pathlist)-1
            return
        if event:
            i_val += 1
            n2 = 0
            if sorted_pathlist[i_val] == 'deleted':
                for h in range(i_val+1, len(sorted_pathlist)):
                    if sorted_pathlist[h] != 'deleted':
                        n2 = h - i_val
                        break
                i_val += n2
                btns[i_val].config(bg="red")
                btns[i_val](background="red", foreground="red",activebackground="red", activeforeground="red")
                btns[i_val-n2-1].config(bg="gray99", fg="purple3", font="Dosis")
                root.after(500, single_click, event, segments, sorted_pathlist[i_val][0])
            else:
                btns[i_val].config(bg="red")
                btns[i_val](background="red", foreground="red",activebackground="red", activeforeground="red")
                root.after(500, single_click, event, segments, sorted_pathlist[i_val][0])
            if prev == 1:
                btns[i_val-1].config(bg="gray99", fg="purple3", font="Dosis")
    prev = 1


c=0
e=0
d=0
z=0
shown = 0
def make_counter():
    global shown,w1,h1,counter_green,counter_red,counter_total,fcounter_red,fcounter_green,fcounter_total
    
    shown = 1
    for widget in frm3.winfo_children():
        widget.destroy()
    datapts = []    #stores (x, y, color) tuple for every click 
    drawpts = []    #stores the dots
    undolist = []   #delets the dots
    undopts = []    #stores the deleted (x, y, color) tuple
    head,tail=os.path.split(undo_path)
  
    head,tail=os.path.split(head)
    head=head+"/segments/scoredData.xlsx"
    wb = openpyxl.load_workbook(head)
    ws = wb.active
    
    for i in range(0, ws.max_row-1):
        if mc_and_dc_list[i][3] == filename_ext:
    
            temp_i = i
            break        

    def clickm(event,a):
        
        x1 = event.x-(w1+h1)/4
        y1=event.y-(w1+h1)/4
        x2=event.x+(w1+h1)/4
        y2=event.y+(w1+h1)/4
        id=canvas.create_oval(x1,y1,x2,y2,fill='',width=2,outline='red')
        
       
        f =open(undo_path,'a')
        f.write('{}  {}  {}  {}  {}  {}  {}\n'.format(x1,y1,x2,y2,0,id,a))
        f.close
        
        if(a==1):
            wb = openpyxl.load_workbook(head)
            ws = wb.active
            c1=ws.cell(temp_i+2,5)
            c1.value+=1
            mc_and_dc_list[temp_i][0]+=1
            counter2.set(c1.value)
            #global e 
            #e=e+1
            c1=ws.cell(temp_i+2, 7)
            c1.value+=1
            mc_and_dc_list[temp_i][2]+=1
            counter3.set(c1.value)
            counter_red.set(counter_red.get() +1)
            counter_total.set(counter_total.get() +1)
            fcounter_red.set(fcounter_red.get() +1)
            fcounter_total.set(fcounter_total.get() +1)
            wb.save(head)
        if(a==2):
            wb = openpyxl.load_workbook(head)
            ws = wb.active
            c1=ws.cell(temp_i+2,6)
            c1.value-=1
            mc_and_dc_list[temp_i][1]-=1
            counter1.set(c1.value)
            #global e 
            #e=e+1
            c1=ws.cell(temp_i+2, 7)
            c1.value-=1
            mc_and_dc_list[temp_i][2]-=1
            counter3.set(c1.value)
            counter_green.set(counter_green.get() -1)
            counter_total.set(counter_total.get() -1)
            fcounter_green.set(fcounter_green.get() -1)
            fcounter_total.set(fcounter_total.get() -1)
            wb.save(head)
        



    def clickd(event,a):
        
        x1 = event.x-(w1+h1)/4
        y1=event.y-(w1+h1)/4
        x2=event.x+(w1+h1)/4
        y2=event.y+(w1+h1)/4
        global d  
        d=d+1
        id=canvas.create_oval(x1,y1,x2,y2,fill='',width=2,outline='green')
        
        f =open(undo_path,'a')
        f.write('{}  {}  {}  {}  {}  {}  {}\n'.format(x1,y1,x2,y2,1,id,a))
        f.close
        
        if(a==1):
            wb = openpyxl.load_workbook(head)
            ws = wb.active
            c1=ws.cell(temp_i+2, 5)
            c1.value-=1
            mc_and_dc_list[temp_i][0]-=1
            counter2.set(c1.value)
            #global e 
            #e=e+1
            c1=ws.cell(temp_i+2, 7)
            c1.value-=1
            mc_and_dc_list[temp_i][2]-=1
            counter3.set(c1.value)
          
            counter_red.set(counter_red.get() -1)
            counter_total.set(counter_total.get() -1)
            fcounter_red.set(fcounter_red.get() -1)
            fcounter_total.set(fcounter_total.get() -1)
            wb.save(head)
        if(a==2):
            wb = openpyxl.load_workbook(head)
            ws = wb.active
            c1=ws.cell(temp_i+2,6)
            c1.value+=1
            mc_and_dc_list[temp_i][1]+=1
            counter1.set(c1.value)
        
            #global e 
            #e=e+1
            c1=ws.cell(temp_i+2, 7)
            c1.value+=1
            mc_and_dc_list[temp_i][2]+=1
            counter3.set(c1.value)
         
            counter_green.set(counter_green.get() +1)
            counter_total.set(counter_total.get() +1)
            fcounter_green.set(fcounter_green.get() +1)
            fcounter_total.set(fcounter_total.get() +1)
            wb.save(head)

    def undo():
        f =open(undo_path,'r')
        f.seek(0, 0)
        
        data = np.genfromtxt(undo_path)
        f.close()
        row,col=np.shape(data)
        
        if row > 1:
            x1,y1,x2,y2,colour,id,a=data[row-1]
            
            canvas.delete(int(id))
            f=open(redo_path,'a')
            f.write('{}  {}  {}  {}  {}  {}  {}\n'.format(x1,y1,x2,y2,colour,id,a))
            f.close
            data=np.delete(data,(row-1),axis=0)
            data=np.delete(data,(0),axis=0)
            f =open(undo_path,'w')
            f.write("x1  y1  x2  y2  colour  id  type\n")
            for part in data:
                f.write('{}  {}  {}  {}  {}  {}  {}\n'.format(part[0],part[1],part[2],part[3],part[4],part[5],part[6]))
            f.close()        
            wb = openpyxl.load_workbook(head)
            ws = wb.active
           
            if colour == 0:
                if ((mc_and_dc_list[temp_i][0]>=0)  and  ( mc_and_dc_list[temp_i][1]>=0)) :
                    if(a==1):
                        wb = openpyxl.load_workbook(head)
                        ws = wb.active
                        c1=ws.cell(temp_i+2,5)
                        c1.value-=1
                        mc_and_dc_list[temp_i][0]-=1
                        counter2.set(c1.value)
                    
                        #global e 
                        #e=e+1
                        c1=ws.cell(temp_i+2, 7)
                        c1.value-=1
                        mc_and_dc_list[temp_i][2]-=1
                        counter3.set(c1.value)
                      
                        counter_red.set(counter_red.get() -1)
                        counter_total.set(counter_total.get() -1)
                        fcounter_red.set(fcounter_red.get() -1)
                        fcounter_total.set(fcounter_total.get() -1)
                        wb.save(head)
                    if(a==2):
                        wb = openpyxl.load_workbook(head)
                        ws = wb.active
                        c1=ws.cell(temp_i+2,6)
                        c1.value+=1
                        mc_and_dc_list[temp_i][1]+=1
                        counter1.set(c1.value)
                      
                        #global e 
                        #e=e+1
                        c1=ws.cell(temp_i+2, 7)
                        c1.value+=1
                        mc_and_dc_list[temp_i][2]+=1
                        counter3.set(c1.value)
                    
                        counter_green.set(counter_green.get() +1)
                        counter_total.set(counter_total.get() +1)
                        fcounter_green.set(fcounter_green.get() +1)
                        fcounter_total.set(fcounter_total.get() +1)
                        wb.save(head)

            elif colour == 1:
                if mc_and_dc_list[temp_i][1]>0 :
                    if(a==1):
                        wb = openpyxl.load_workbook(head)
                        ws = wb.active
                        c1=ws.cell(temp_i+2, 5)
                        c1.value+=1
                        mc_and_dc_list[temp_i][0]+=1
                        counter2.set(c1.value)
                        #global e 
                        #e=e+1
                        c1=ws.cell(temp_i+2, 7)
                        c1.value+=1
                        mc_and_dc_list[temp_i][2]+=1
                        counter3.set(c1.value)
                       
                        counter_red.set(counter_red.get() +1)
                        counter_total.set(counter_total.get() +1)
                        fcounter_red.set(fcounter_red.get() +1)
                        fcounter_total.set(fcounter_total.get() +1)
                        wb.save(head)
                    if(a==2):
                        wb = openpyxl.load_workbook(head)
                        ws = wb.active
                        c1=ws.cell(temp_i+2,6)
                        c1.value-=1
                        mc_and_dc_list[temp_i][1]-=1
                        counter1.set(c1.value)
                     
                        #global e 
                        #e=e+1
                        c1=ws.cell(temp_i+2, 7)
                        c1.value-=1
                        mc_and_dc_list[temp_i][2]-=1
                        counter3.set(c1.value)
                       
                        counter_green.set(counter_green.get() -1)
                        counter_total.set(counter_total.get() -1)
                        fcounter_green.set(fcounter_green.get() -1)
                        fcounter_total.set(fcounter_total.get() -1)
                        wb.save(head)
            

    def redo():
        f =open(redo_path,'r')
        f.seek(0, 0)
        data = np.genfromtxt(redo_path)
        f.close()
        row,col=np.shape(data)
        if row > 1:
            x1,y1,x2,y2,colour,id,a=data[row-1]
           
            f=open(undo_path,'a')
            if(colour==1):
                id=canvas.create_oval(x1,y1,x2,y2,fill='',width=2,outline='green')
            else:
                id=canvas.create_oval(x1,y1,x2,y2,fill='',width=2,outline='red')
            f.write('{}  {}  {}  {}  {}  {}  {}\n'.format(x1,y1,x2,y2,colour,id,a))
            f.close
            data=np.delete(data,(row-1),axis=0)
            data=np.delete(data,(0),axis=0)
            f =open(redo_path,'w')
            f.write("x1  y1  x2  y2  colour  id  type\n")
            for part in data:
                f.write('{}  {}  {}  {}  {}  {}  {}\n'.format(part[0],part[1],part[2],part[3],part[4],part[5],part[6]))
            f.close()
            

            
            wb = openpyxl.load_workbook(head)
            ws = wb.active
         
            if colour == 0:
                if ((mc_and_dc_list[temp_i][0]>=0)  and  ( mc_and_dc_list[temp_i][1]>=0)) :
                    if(a==1):
                        wb = openpyxl.load_workbook(head)
                        ws = wb.active
                        c1=ws.cell(temp_i+2,5)
                        c1.value+=1
                        mc_and_dc_list[temp_i][0]+=1
                        counter2.set(c1.value)
                       
                        #global e 
                        #e=e+1
                        c1=ws.cell(temp_i+2, 7)
                        c1.value+=1
                        mc_and_dc_list[temp_i][2]+=1
                        counter3.set(c1.value)
                      
                        counter_red.set(counter_red.get() +1)
                        counter_total.set(counter_total.get() +1)
                        fcounter_red.set(fcounter_red.get() +1)
                        fcounter_total.set(fcounter_total.get() +1)
                        wb.save(head)
                    if(a==2):
                        wb = openpyxl.load_workbook(head)
                        ws = wb.active
                        c1=ws.cell(temp_i+2,6)
                        c1.value-=1
                        mc_and_dc_list[temp_i][1]-=1
                        counter1.set(c1.value)
                    
                        #global e 
                        #e=e+1
                        c1=ws.cell(temp_i+2, 7)
                        c1.value-=1
                        mc_and_dc_list[temp_i][2]-=1
                        counter3.set(c1.value)
                      
                        counter_green.set(counter_green.get() -1)
                        counter_total.set(counter_total.get() -1)
                        fcounter_green.set(fcounter_green.get() -1)
                        fcounter_total.set(fcounter_total.get() -1)
                        wb.save(head)

            elif colour == 1:
                if mc_and_dc_list[temp_i][1]>=0 :
                    if(a==1):
                        wb = openpyxl.load_workbook(head)
                        ws = wb.active
                        c1=ws.cell(temp_i+2, 5)
                        c1.value-=1
                        mc_and_dc_list[temp_i][0]-=1
                        counter2.set(c1.value)
                        #global e 
                        #e=e+1
                        c1=ws.cell(temp_i+2, 7)
                        c1.value-=1
                        mc_and_dc_list[temp_i][2]-=1
                        counter3.set(c1.value)
                     
                        counter_red.set(counter_red.get() -1)
                        counter_total.set(counter_total.get() -1)
                        fcounter_red.set(fcounter_red.get() -1)
                        fcounter_total.set(fcounter_total.get() -1)
                        wb.save(head)
                    if(a==2):
                        wb = openpyxl.load_workbook(head)
                        ws = wb.active
                        c1=ws.cell(temp_i+2,6)
                        c1.value+=1
                        mc_and_dc_list[temp_i][1]+=1
                        counter1.set(c1.value)
                       
                        #global e 
                        #e=e+1
                        c1=ws.cell(temp_i+2, 7)
                        c1.value+=1
                        mc_and_dc_list[temp_i][2]+=1
                        counter3.set(c1.value)
                       
                        counter_green.set(counter_green.get() +1)
                        counter_total.set(counter_total.get() +1)
                        fcounter_green.set(fcounter_green.get() +1)
                        fcounter_total.set(fcounter_total.get() +1)
                        wb.save(head)
        
    def retrieve():
      
        if os.path.isfile(head):
            pass
        else:
            return
        f =open(undo_path,'r')
        f.seek(0, 0)
        data = np.genfromtxt(undo_path)
        f.close()
        row,col=np.shape(data)
        data=np.delete(data,(0),axis=0)
        f =open(undo_path,'w')
        f.write("x1  y1  x2  y2  colour  id  type\n")
        for z in data:
            x1,y1,x2,y2,colour,id,a=z
           
            if(colour==1):
                id=canvas.create_oval(x1,y1,x2,y2,fill='',width=2,outline='green')
            else:
                id=canvas.create_oval(x1,y1,x2,y2,fill='',width=2,outline='red')
            f.write('{}  {}  {}  {}  {}  {}  {}\n'.format(x1,y1,x2,y2,colour,id,a))
        f.close()
        f =open(redo_path,'w')
        f.write("x1  y1  x2  y2  colour  id  type\n")
        f.close()
        wb = openpyxl.load_workbook(head)
        ws = wb.active
        c1=ws.cell(temp_i+2,5)
        counter2.set(c1.value)
        c1=ws.cell(temp_i+2,6)
        counter1.set(c1.value)
        c1=ws.cell(temp_i+2,7)
        counter3.set(c1.value)
        row=ws.max_row
        col=ws.max_column
        for i in range(2, row+1 ):
            for j in range(col-3,col+1):
                c1=ws.cell(i, j)
        wb.save(head)


        #canvas.update()
        #ps = canvas.postscript(file = file_name, colormode = 'color')
    
    
   
    
    def mono():
        
        canvas.bind("<Button-1>",lambda event:clickm(event,2))
        canvas.bind("<Button-3>",lambda event:clickd(event,2))
    
    def dice():
        canvas.bind("<Button-1>",lambda event:clickm(event,1))
        canvas.bind("<Button-3>",lambda event:clickd(event,1))


    canvas.create_image(0, 0, image = photo1, anchor = NW)   
    
    #canvas.bind("Key",clicka)
    
    counter1=IntVar()
    counter1.set(mc_and_dc_list[temp_i][1])
    #label0=Label(frm3,bg="white",text="Click to add a DC\nDouble-click to\ndiscard a DC",width=15)
    #label0.pack(fill=X, pady=10)
    label1 = Button(frm3,bg="white",text="Monocentric",width=15,command=mono)
    
    label1.pack(fill=X, pady=10)
    label2= Label(frm3,bg="#80c1ff",textvariable=counter1,width=15)
    label2.pack(fill=X, pady=10)

    counter2=IntVar()
    counter2.set(mc_and_dc_list[temp_i][0])
    label3=Button(frm3,bg="white",text="Dicentric",width=15,command=dice)
    label3.pack(fill=X, pady=10)  
    label4= Label(frm3,bg="#80c1ff",textvariable=counter2,width=15)
    label4.pack(fill=X, pady=10)

    #savebtn = Tkinter.Button()
#    counter4=IntVar()
#    counter4.set(0)
#    label5=Label(frm3,bg="white",text="Acentric",width=15)
#    label5.pack(fill=X, pady=10)
#    label6= Label(frm3,bg="#80c1ff",textvariable=counter4,width=15)
#    label6.pack(fill=X, pady=10)

    counter3=IntVar()
    counter3.set(mc_and_dc_list[temp_i][2])
    label7=Label(frm3,bg="white",text="Total Segments",width=15)
    label7.pack(fill=X, pady=10)
    label8= Label(frm3,bg="#80c1ff",textvariable=counter3,width=15)
    label8.pack(fill=X, pady=10)
    button10=Button(frm3,bg="white",text="Undo",width=15, command=undo)
    button10.pack(fill=X,pady=10)
    button11=Button(frm3,bg="white",text="Redo",width=15, command=redo)
    button11.pack(fill=X,pady=10)
    button9=Button(frm3,bg="white",text="Retrieve",width=15,command=retrieve)
    button9.pack(fill=X,pady=10)
    retrieve()
    def manual_file():
        soft_data=openpyxl.Workbook()
        soft_sheet = soft_data.active
        soft_sheet.title="manual_scoring"
        c1 = soft_sheet['A1'] 
        c1.value = "Metaphase_ID" 
        c2 = soft_sheet['B1'] 
        c2.value = "Monocentric"
        c3 = soft_sheet['C1'] 
        c3.value = "Dicentric" 
        c4 = soft_sheet['D1'] 
        c4.value = "Acentic"
        c5 = soft_sheet['E1'] 
        c5.value = "Total"
        c12 = soft_sheet['A2'] 
        c12.value = file_name1
        c22 = soft_sheet['B2'] 
        c22.value = c
        c32 = soft_sheet['C2'] 
        c32.value = d
        c42 = soft_sheet['D2'] 
        c42.value = z
        c52 = soft_sheet['E2'] 
        c52.value = e
        soft_data.save("manual_scoring%s.xlsx"%time.time())



def display_metaphase():
    global fcounter_red,fcounter_green,fcounter_total
    images = []
    buttons = []
    row_number = 0
    iteration = 0
    for widget in scframe.interior.winfo_children():
        widget.destroy()
    canvas.delete("all")
    min_score=threshold_entry.get()
   
    files = 0
    fred = 0
    fgreen = 0
    ftotal = 0
    x = originalPath[0][0].split("/")
    original_path = "/"
    for i in range(len(x)-3):
        original_path = original_path + str(x[i+1]) + "/"
    real_path = original_path + str(x[-2])
    sorted_pathlist=sorted(originalPath, key=lambda x:x[1][0])
    
    temp=0
    flag=0
    for i in range(len(sorted_pathlist)):

        k = sorted_pathlist[i][0]
        v = sorted_pathlist[i][1]
        if int(v[0])<=int(min_score):
            
        # break
            
            fred = fred + mc_and_dc_list[-i-1][0]
            fgreen = fgreen + mc_and_dc_list[-i-1][1]
            ftotal = ftotal + mc_and_dc_list[-i-1][2]
            files = files + 1
            for x in xrange(len(v)):
                v[x] = str(v[x])
            text = k
            segments = v[1]
         
            if temp != v[0]:
                iteration=0
                row_number+=1
#                setofimgs=set()
                texttodisp = 'Rank: ' + v[0]
                w=Tkinter.Label(scframe.interior, text=texttodisp, height=1, width=30, bg="gray", fg="black", relief=Tkinter.FLAT)
                w.grid(row= row_number, columnspan=2)
                row_number+=1
#                tkvar = StringVar(root) 
            
            image = Image.open(text)
            image = image.resize((100,100), Image.ANTIALIAS)
            photo = ImageTk.PhotoImage(image)
            btn = Tkinter.Button(scframe.interior,height=2, width=20, relief=Tkinter.FLAT, bg="gray99", fg="purple3",font="Dosis")
            btn.config(image=photo,width="100",height="100")
            btn.image=photo
            btn.grid(row=row_number, column=iteration)
            images.append(photo)
            buttons.append(btn)
            if iteration==1:
                row_number+=1
            iteration = 1-iteration
#            setofimgs.add(btn)
#            popupMenu = OptionMenu(root, tkvar, *setofimgs)
#            popupMenu.pack()
#            tkvar.trace('w', change_dropdown)
            btn.bind('<Button-1>', lambda event, segments = segments, text=text :test(event, sorted_pathlist,buttons,segments,text,0,1))
            btn.bind('<Double-Button-1>', lambda event, segments = segments, text=text :test(event,sorted_pathlist,buttons,segments,text,0,101))
            temp=v[0]
                 

    ffiles=IntVar()
    ffiles.set(files)
    fcounter_red=IntVar()
    fcounter_green=IntVar()
    fcounter_total=IntVar()
    fcounter_red.set(fred)
    fcounter_green.set(fgreen)
    fcounter_total.set(ftotal)


    labels01=Label(f3, text="Files", bg="gray", fg="black", width = 25).pack(fill=X, pady=5)
    labels02=Label(f3, textvariable=ffiles, bg="white", fg="black", width = 25).pack(fill=X, pady=5)

    labels11=Label(f3, text="Dicentric", bg="gray", fg="black", width = 25).pack(fill=X, pady=5)
    labels12=Label(f3, textvariable=fcounter_red, bg="white", fg="black", width = 25).pack(fill=X, pady=5)

    labels21=Label(f3, text="Monocentric", bg="gray", fg="black", width = 25).pack(fill=X, pady=5)
    labels22=Label(f3, textvariable=fcounter_green, bg="white", fg="black", width = 25).pack(fill=X, pady=5)
        
    labels31=Label(f3, text="Total Chromosome ", bg="gray", fg="black", width = 25).pack(fill=X, pady=5)
    labels32=Label(f3, textvariable=fcounter_total, bg="white", fg="black", width = 25).pack(fill=X, pady=5)

    New_path = real_path + '/segments/color.xlsx'
    if os.path.isfile(New_path):
        soft_dat = openpyxl.load_workbook(New_path)
        soft_she = soft_dat.active
        new_way = 2
        for d in range(soft_she.max_row):
            fin1 = soft_she.cell(new_way,1)
            new_way = new_way + 1
            g = fin1.value
            buttons[g].config(bg="red")
               

    prevBtn = Button(f3, text='<', bg='gray99', fg='black')
    prevBtn.place(relx=0.80, rely=0.99, anchor=SE)
    prevBtn.bind('<Button-1>', lambda event, segments = segments, text=text :test(event,sorted_pathlist,buttons,segments,text,1, 1))
    nextBtn = Button(f3, text='>', bg='gray99', fg='black')
    nextBtn.place(relx=0.90, rely=0.99, anchor=SE)
    nextBtn.bind('<Button-1>', lambda event, segments = segments, text=text :test(event,sorted_pathlist,buttons,segments,text, 2, 1))


def display_rank():
    plt.bar(range(1,len(rankPlot)+1),rankPlot)
    plt.title('Rank Statistics')
    plt.ylabel('Number of metaphases')
    plt.xlabel('Rank of metaphase')
    plt.grid(True,color='k')
    plt.show()
    plt.close()

def image_path():           # For Input

    global file_name_list,f2,f3, root, scframe, frm1,frm2,frm3,frm4,frm5,frm6, canvas, canvas2
    root, f2, f3, scframe, canvas, f4,  canvas2, frm1,frm2,frm3,frm4,frm5,frm6 = make_gui_before()
    root.mainloop()
    file_name_list = []

    if selected_file != "":
        file_name_list.append(selected_file)
    if selected_folder != "":
        for file in os.listdir(selected_folder):
            if file.endswith(extension):
                file_name_list.append(os.path.join(selected_folder, file))

    file_name_list.sort()
    if not file_name_list:
        print "Select appropriate folder"
        sys.exit()
    else :

        global progress,mc_and_dc_list
        
    
        
        progress= ttk.Progressbar(f2, orient = 'horizontal', maximum = len(file_name_list), variable=downloaded, mode = 'determinate')
        progress.pack(fill=BOTH)
        start = ttk.Button(f2,text='Run ranking algorithm',command= lambda: Threader())
        start.pack(fill=BOTH)
        root.mainloop()
        b2=Button(f3, text = "Show Rank Statistics", command=display_rank)
        b2.pack(fill=X, pady = 30)
        
        global threshold_entry
        threshold_entry = Entry(f3, bd =5)
        threshold_entry.delete(0,'end')
        threshold_entry.insert(1, 10)
        threshold_entry.pack(fill = X)
        b = Button(f3, text = "Show metaphases with rank upto:",command=display_metaphase)
        b.pack(fill=X)
        global image_name_entry
        Label(f3, text="Selected Image : ", bg="gray", fg="black", width = 25).pack(fill=X, pady=10)
        image_name_entry = Entry(f3, bd =5)
        image_name_entry.pack(fill = X)
        global total_segments
        Label(f3, text="Number of segments", bg="gray", fg="black", width = 25).pack(fill=X, pady=10)
        total_segments = Entry(f3, bd =5)
        total_segments.pack(fill = X)
        global metaphase_rank
        Label(f3, text="Rank", bg="gray", fg="black", width = 25).pack(fill=X, pady=10)
        metaphase_rank = Entry(f3, bd =5)
        metaphase_rank.pack(fill = X)
        global b4
        b4=Button(f3, text = "Detect DCs")
        b4.pack(fill=X, pady = 30)
        b4.bind('<Button-1>', lambda event, segments = '', text='' :detect_dc(event,segments,text,1))
        #text="File"+str(len(mc_and_dc_list))
        
        
        #Label(f3, text=text, bg="gray", fg="black", width = 25).pack(fill=X, pady=10)
        root.mainloop()

        return file_name_list, f2, root


if __name__ == '__main__':
    red_cross = Image.open("Red-Cross-PNG-File.png")

    file_name_list,f2,root = image_path()
